"""
Flask blueprint for /api/credit-default/*.

Endpoints:
  GET  /api/credit-default/dashboard       Full panel + scored ratings + summary
  GET  /api/credit-default/table           Compact list-of-rows for the table view
  GET  /api/credit-default/country/<iso3>  Drilldown for one country
  GET  /api/credit-default/methodology     Returns weights + rating buckets
  POST /api/credit-default/refresh         Force re-fetch + re-score
"""

from __future__ import annotations

import io
import time

from flask import Blueprint, jsonify, request, send_file

from backend.credit_default import rating_model, service


credit_default_bp = Blueprint('credit_default', __name__)


def _parse_cadence_args():
    """Parse ?cadence=annual|quarterly and ?horizon=N from query args.

    Annual horizon ∈ {1, 3, 5} (years); quarterly horizon ∈ {4, 12, 20}
    (quarters). Falls back to annual / 1y on any garbage input.
    """
    cadence = (request.args.get('cadence') or 'annual').lower()
    if cadence not in ('annual', 'quarterly'):
        cadence = 'annual'
    try:
        horizon = int(request.args.get('horizon', 1))
    except (TypeError, ValueError):
        horizon = 1
    if cadence == 'quarterly':
        if horizon not in (4, 12, 20):
            horizon = 4
    else:
        if horizon not in (1, 3, 5):
            horizon = 1
    return cadence, horizon


@credit_default_bp.route('/dashboard')
def dashboard():
    cadence, horizon = _parse_cadence_args()
    return jsonify(service.get_dashboard(cadence=cadence, horizon=horizon))


@credit_default_bp.route('/table')
def table():
    cadence, horizon = _parse_cadence_args()
    return jsonify({
        'rows': service.get_table_rows(cadence=cadence, horizon=horizon),
        'as_of': service.get_dashboard(cadence=cadence, horizon=horizon).get('as_of'),
        'cadence': cadence,
        'horizon': horizon,
    })


@credit_default_bp.route('/country/<iso3>')
def country(iso3: str):
    cadence, horizon = _parse_cadence_args()
    c = service.get_country(iso3, cadence=cadence, horizon=horizon)
    if not c:
        return jsonify({'error': f'country {iso3} not found'}), 404
    return jsonify(c)


@credit_default_bp.route('/country/<iso3>/history')
def country_history(iso3: str):
    cadence, horizon = _parse_cadence_args()
    h = service.get_country_history(
        iso3, horizon_years=horizon, cadence=cadence,
    )
    if h is None:
        return jsonify({'error': f'history unavailable for {iso3}'}), 404
    return jsonify(h)


@credit_default_bp.route('/methodology')
def methodology():
    return jsonify({
        'weights': rating_model.WEIGHTS,
        'higher_is_worse': rating_model.HIGHER_IS_WORSE,
        'rating_buckets': [
            {'max_score': r[0], 'sp': r[1], 'moodys': r[2],
             'pd_1y': r[3], 'pd_3y': r[4], 'pd_5y': r[5]}
            for r in rating_model.RATING_BUCKETS
        ],
        'z_clip': rating_model.Z_CLIP,
    })


@credit_default_bp.route('/refresh', methods=['POST'])
def refresh():
    data = service.get_dashboard(force_refresh=True)
    return jsonify({
        'as_of': data.get('as_of'),
        'country_count': len((data.get('countries') or {})),
    })


@credit_default_bp.route('/export')
def export_xlsx():
    """Excel export for AIG analysts: every row in the table view, plus the
    full indicator panel as a second sheet.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    rows = service.get_table_rows()
    dashboard = service.get_dashboard()
    countries = dashboard.get('countries') or {}
    indicators_meta = dashboard.get('indicators') or {}

    wb = Workbook()

    # Sheet 1 — ratings + PD + agency comparison (table view)
    ws = wb.active
    ws.title = 'Ratings & PD'
    cols = [
        ('iso3', 'ISO3'), ('name', 'Country'), ('region', 'Region'),
        ('source', 'Source'),
        ('pm_notch', 'PM Rating'), ('sp_equiv', 'S&P equiv'),
        ('moodys_equiv', "Moody's equiv"),
        ('composite_pm_notch', 'Composite'),
        ('agency_sp', 'S&P'), ('agency_moodys', "Moody's"), ('agency_fitch', 'Fitch'),
        ('notch_delta_sp', 'Δ vs S&P'),
        ('pd_1y', 'PD 1y'), ('pd_3y', 'PD 3y'), ('pd_5y', 'PD 5y'),
        ('composite_pd_1y', 'Composite PD 1y'),
        ('shadow_debt_gap_pp', 'Shadow gap (pp)'),
        ('risk_tier', 'Risk tier'),
    ]
    ws.append([h for _, h in cols])
    for r in rows:
        ws.append([r.get(k) for k, _ in cols])

    # Sheet 2 — full indicator panel
    ws2 = wb.create_sheet('Indicators')
    indicator_keys = list(indicators_meta.keys())
    ws2.append(['ISO3', 'Country', 'Region'] + [indicators_meta[k]['label'] for k in indicator_keys])
    for iso3, c in countries.items():
        ind = c.get('indicators') or {}
        ws2.append([iso3, c.get('name'), c.get('region')] + [ind.get(k) for k in indicator_keys])

    # Sheet 3 — methodology
    ws3 = wb.create_sheet('Methodology')
    model = dashboard.get('model') or {}
    ws3.append(['Model', model.get('name', '')])
    ws3.append(['Version', model.get('version', '')])
    ws3.append(['Estimator', model.get('estimator', '')])
    ws3.append(['Method', model.get('method', '')])
    ws3.append(['Scale', model.get('scale', '')])
    ws3.append([])
    ws3.append(['Indicator', 'Scaffold weight', 'Higher is worse?', 'Fitted coef'])
    weights = model.get('weights') or {}
    higher = model.get('higher_is_worse') or {}
    fitted = model.get('fitted_coefficients') or {}
    for k in weights.keys():
        ws3.append([k, weights[k], higher.get(k), fitted.get(k)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'parra-credit-default_{time.strftime("%Y%m%d")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname,
    )
