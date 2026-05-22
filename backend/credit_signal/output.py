"""Write Excel (multi-sheet, colour-coded) and CSV outputs for the matrix."""

import os

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ISO3 -> business region. Extend as needed for additional country coverage.
REGIONS = {
    'USA': 'North America', 'CAN': 'North America', 'MEX': 'North America',
    'BRA': 'LATAM', 'ARG': 'LATAM', 'COL': 'LATAM', 'CHL': 'LATAM',
    'PER': 'LATAM', 'VEN': 'LATAM',
    'GBR': 'Europe', 'FRA': 'Europe', 'DEU': 'Europe', 'ITA': 'Europe',
    'ESP': 'Europe', 'NLD': 'Europe', 'POL': 'Europe', 'CHE': 'Europe',
    'SWE': 'Europe', 'NOR': 'Europe',
    'RUS': 'CEEMEA', 'TUR': 'CEEMEA', 'SAU': 'CEEMEA', 'ARE': 'CEEMEA',
    'ZAF': 'CEEMEA', 'NGA': 'CEEMEA', 'EGY': 'CEEMEA', 'QAT': 'CEEMEA',
    'KAZ': 'CEEMEA',
    'CHN': 'APAC', 'JPN': 'APAC', 'KOR': 'APAC', 'IND': 'APAC',
    'IDN': 'APAC', 'AUS': 'APAC', 'THA': 'APAC', 'MYS': 'APAC',
    'VNM': 'APAC', 'PHL': 'APAC',
}

# Signal -> (cell fill, font colour). Traffic-light palette.
_SIGNAL_FILL = {
    'AVOID':     ('C0392B', 'FFFFFF'),  # red
    'CAUTION':   ('E67E22', 'FFFFFF'),  # amber
    'NEUTRAL':   ('BDC3C7', '000000'),  # grey
    'STRATEGIC': ('27AE60', 'FFFFFF'),  # green
    'NA':        ('FFFFFF', '000000'),
}


def write_outputs(sovereign_df, sector_df, methodology, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, 'credit_signal_matrix.xlsx')
    csv_long_path = os.path.join(out_dir, 'credit_signal_long.csv')
    csv_sov_path = os.path.join(out_dir, 'credit_signal_sovereign.csv')

    sov_sorted = sovereign_df.sort_values('adjusted_pd', ascending=False)

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as w:
        sov_sorted.to_excel(w, sheet_name='Sovereign', index=False)
        wide_signal = wide_pd = None
        if not sector_df.empty:
            sector_sorted = sector_df.sort_values(
                ['sector_pd', 'book_weight_pct'], ascending=[False, False])
            sector_sorted.to_excel(w, sheet_name='Country_x_Sector_Long',
                                   index=False)
            wide_signal = sector_df.pivot_table(
                index=['iso3', 'country'], columns='sector',
                values='signal', aggfunc='first')
            wide_signal.to_excel(w, sheet_name='Signal_Matrix_Wide')
            wide_pd = sector_df.pivot_table(
                index=['iso3', 'country'], columns='sector',
                values='sector_pd', aggfunc='first')
            wide_pd.to_excel(w, sheet_name='PD_Matrix_Wide')
        regional = _regional_rollup(sovereign_df)
        regional.to_excel(w, sheet_name='Regional_Summary', index=False)
        pd.DataFrame(list(methodology.items()),
                     columns=['Item', 'Value']).to_excel(
            w, sheet_name='Methodology', index=False)

        # --- colour-coding (after data is written, same writer/workbook) ---
        wb = w.book
        _style_signal_column(wb['Sovereign'], sov_sorted)
        if wide_signal is not None:
            _style_signal_grid(wb['Signal_Matrix_Wide'], wide_signal)
            _style_signal_column(wb['Country_x_Sector_Long'], sector_sorted)
            _style_pd_heatmap(wb['PD_Matrix_Wide'], wide_pd)
        _autofit(wb['Sovereign'])
        _autofit(wb['Regional_Summary'])

    sector_df.to_csv(csv_long_path, index=False)
    sov_sorted.to_csv(csv_sov_path, index=False)
    return {'xlsx': xlsx_path,
            'sovereign_csv': csv_sov_path,
            'matrix_csv': csv_long_path}


def _style_signal_column(ws, df):
    """Colour the 'signal' column of a long-form sheet (header in row 1)."""
    if 'signal' not in df.columns:
        return
    col_idx = list(df.columns).index('signal') + 1  # 1-based, no index col
    for r, val in enumerate(df['signal'].values, start=2):
        bg, fg = _SIGNAL_FILL.get(val, _SIGNAL_FILL['NA'])
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.font = Font(color=fg, bold=True)
        cell.alignment = Alignment(horizontal='center')


def _style_signal_grid(ws, wide):
    """Colour every sector cell of the wide signal grid.

    Layout written by pandas: cols A,B = iso3,country; sector values start
    at column C (3); data starts at row 2.
    """
    n_index_cols = 2  # iso3, country
    for r in range(len(wide)):
        for c in range(wide.shape[1]):
            val = wide.iat[r, c]
            bg, fg = _SIGNAL_FILL.get(val, _SIGNAL_FILL['NA'])
            cell = ws.cell(row=r + 2, column=c + 1 + n_index_cols)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.font = Font(color=fg, bold=True)
            cell.alignment = Alignment(horizontal='center')
    _autofit(ws)


def _style_pd_heatmap(ws, wide):
    """Apply a green->red colour scale across the PD value cells."""
    n_index_cols = 2
    first_col = get_column_letter(n_index_cols + 1)
    last_col = get_column_letter(n_index_cols + wide.shape[1])
    last_row = len(wide) + 1
    rng = f'{first_col}2:{last_col}{last_row}'
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type='min', start_color='27AE60',
        mid_type='percentile', mid_value=50, mid_color='F1C40F',
        end_type='max', end_color='C0392B'))
    for row in ws.iter_rows(min_row=2, min_col=n_index_cols + 1):
        for cell in row:
            cell.number_format = '0.0000'
    _autofit(ws)


def _autofit(ws, max_width=28):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None),
                     default=8)
        ws.column_dimensions[col[0].column_letter].width = min(
            max_width, max(8, length + 1))


def _regional_rollup(sov):
    sov = sov.copy()
    sov['region'] = sov['iso3'].map(REGIONS).fillna('Other')
    grouped = sov.groupby('region').agg(
        countries=('iso3', 'count'),
        median_baseline_pd=('baseline_pd', 'median'),
        median_adjusted_pd=('adjusted_pd', 'median'),
        median_pd_change_pct=('pd_change_pct', 'median'),
        avoid=('signal', lambda s: int((s == 'AVOID').sum())),
        caution=('signal', lambda s: int((s == 'CAUTION').sum())),
        neutral=('signal', lambda s: int((s == 'NEUTRAL').sum())),
        strategic=('signal', lambda s: int((s == 'STRATEGIC').sum())),
    ).reset_index()
    return grouped
