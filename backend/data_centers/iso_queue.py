"""
ISO interconnection-queue ingestor.

Pulls generator interconnection requests from PJM (JSON API) and ERCOT
(public XLSX), aggregates by county / state, maps to the CBRE metros we
track, and emits a grid-pressure snapshot per metro:

    metro           queue_mw_total   dc_named_mw     top_projects
    Northern VA     12,400           1,800           [...]
    DFW             8,200            900             [...]

`queue_mw_total` is the entire pending generation queue in counties
mapped to the metro — a proxy for grid pressure regardless of whether
each request is data-center-related. `dc_named_mw` is the subset whose
project name contains data-center keywords ("data", "compute", "AI",
known hyperscaler names) — these are typically the behind-the-meter or
co-located generation tied to a specific DC build.

ERCOT's XLSX URL changes monthly; we discover the latest by scraping
their public landing page. PJM has a stable JSON endpoint.

The admin surfaces a preview panel with the per-metro snapshot and a
CSV download. Nothing auto-merges into the markets CSV — the operator
reviews and updates planned_mw manually if desired.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any

import requests

logger = logging.getLogger(__name__)

USER_AGENT = (
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) '
    'Chrome/124.0.0.0 Safari/537.36'
)

# ── PJM ─────────────────────────────────────────────────────────────────
# PJM publishes its New Services Queue as JSON. The endpoint is public.
PJM_QUEUE_URL = 'https://services.pjm.com/PJMPlanningApi/api/Queue'

# ── ERCOT ───────────────────────────────────────────────────────────────
# Generator Interconnection Status report — XLSX path changes monthly,
# so we first hit the resource page and find the latest .xlsx link.
ERCOT_RESOURCE_PAGE = 'https://www.ercot.com/gridinfo/resource'
ERCOT_XLSX_PATTERN = re.compile(
    r'https?://[^\s"\']*ERCOT[-_]?Generation[-_]?Interconnection[-_]?Status[^\s"\']*\.xlsx',
    re.IGNORECASE,
)

# ── MISO ────────────────────────────────────────────────────────────────
# MISO publishes the active Generator Interconnection Queue as a CSV
# export. Endpoint is documented in their public planning portal.
MISO_QUEUE_URL = 'https://www.misoenergy.org/api/giqueue/getexportgenerationqueue?queueType=current'

# ── CAISO ───────────────────────────────────────────────────────────────
# CAISO publishes a "Public Queue Report" XLSX; the URL drifts so we
# scrape the landing page to find it.
CAISO_RESOURCE_PAGE = 'https://www.caiso.com/planning/Pages/InterconnectionQueue'
CAISO_XLSX_PATTERN = re.compile(
    r'https?://[^\s"\']*PublicQueueReport[^\s"\']*\.xlsx',
    re.IGNORECASE,
)

# ── County → metro mapping ──────────────────────────────────────────────
# Hand-curated list of US counties to the 27 CBRE metros we track. A queue
# request in a listed county counts toward that metro; unlisted counties
# bucket to "Other".
COUNTY_TO_METRO = {
    # Northern Virginia
    ('VA', 'loudoun'):        'Northern Virginia',
    ('VA', 'fairfax'):        'Northern Virginia',
    ('VA', 'prince william'): 'Northern Virginia',
    ('VA', 'manassas'):       'Northern Virginia',
    ('VA', 'arlington'):      'Northern Virginia',
    ('VA', 'fauquier'):       'Northern Virginia',
    # Central Virginia / Richmond
    ('VA', 'henrico'):        'Central Virginia-Richmond',
    ('VA', 'chesterfield'):   'Central Virginia-Richmond',
    ('VA', 'richmond'):       'Central Virginia-Richmond',
    ('VA', 'mecklenburg'):    'Central Virginia-Richmond',
    # NY-NJ
    ('NJ', 'hudson'):         'New York-New Jersey',
    ('NJ', 'bergen'):         'New York-New Jersey',
    ('NJ', 'middlesex'):      'New York-New Jersey',
    ('NJ', 'somerset'):       'New York-New Jersey',
    ('NJ', 'mercer'):         'New York-New Jersey',
    ('NJ', 'morris'):         'New York-New Jersey',
    ('NY', 'westchester'):    'New York-New Jersey',
    ('NY', 'queens'):         'New York-New Jersey',
    # Chicago
    ('IL', 'cook'):           'Chicago',
    ('IL', 'dupage'):         'Chicago',
    ('IL', 'kane'):           'Chicago',
    ('IL', 'will'):           'Chicago',
    ('IL', 'lake'):           'Chicago',
    # Columbus
    ('OH', 'franklin'):       'Columbus',
    ('OH', 'licking'):        'Columbus',
    ('OH', 'delaware'):       'Columbus',
    # Atlanta
    ('GA', 'fulton'):         'Atlanta',
    ('GA', 'cobb'):           'Atlanta',
    ('GA', 'dekalb'):         'Atlanta',
    ('GA', 'gwinnett'):       'Atlanta',
    ('GA', 'douglas'):        'Atlanta',
    ('GA', 'clayton'):        'Atlanta',
    # Charlotte
    ('NC', 'mecklenburg'):    'Charlotte',
    ('NC', 'gaston'):         'Charlotte',
    # Phoenix
    ('AZ', 'maricopa'):       'Phoenix',
    ('AZ', 'pinal'):          'Phoenix',
    # Silicon Valley
    ('CA', 'santa clara'):    'Silicon Valley',
    ('CA', 'alameda'):        'Silicon Valley',
    ('CA', 'san mateo'):      'Silicon Valley',
    # Seattle / Hillsboro
    ('WA', 'king'):           'Seattle',
    ('WA', 'grant'):          'Seattle',
    ('OR', 'washington'):     'Hillsboro-Portland',
    ('OR', 'multnomah'):      'Hillsboro-Portland',
    # DFW / Houston / Austin (ERCOT)
    ('TX', 'dallas'):         'Dallas-Fort Worth',
    ('TX', 'tarrant'):        'Dallas-Fort Worth',
    ('TX', 'collin'):         'Dallas-Fort Worth',
    ('TX', 'denton'):         'Dallas-Fort Worth',
    ('TX', 'ellis'):          'Dallas-Fort Worth',
    ('TX', 'taylor'):         'Dallas-Fort Worth',          # Abilene
    ('TX', 'harris'):         'Houston',
    ('TX', 'fort bend'):      'Houston',
    ('TX', 'travis'):         'Austin-San Antonio',
    ('TX', 'williamson'):     'Austin-San Antonio',
    ('TX', 'bexar'):          'Austin-San Antonio',
    # Reno-Las Vegas
    ('NV', 'clark'):          'Reno-Las Vegas',
    ('NV', 'washoe'):         'Reno-Las Vegas',
    ('NV', 'storey'):         'Reno-Las Vegas',
    # Salt Lake / Boise / Cheyenne / Memphis
    ('UT', 'salt lake'):      'Salt Lake City',
    ('UT', 'utah'):           'Salt Lake City',
    ('ID', 'ada'):            'Boise',
    ('ID', 'canyon'):         'Boise',
    ('WY', 'laramie'):        'Cheyenne',
    ('TN', 'shelby'):         'Memphis',
    ('MS', 'desoto'):         'Memphis',
    # Iowa / Omaha / Kansas City / Indianapolis / Nashville / Minneapolis / Denver / Boston
    ('IA', 'pottawattamie'):  'Omaha-Council Bluffs',
    ('NE', 'douglas'):        'Omaha-Council Bluffs',
    ('IA', 'polk'):           'Des Moines',
    ('IA', 'dallas'):         'Des Moines',
    ('MO', 'jackson'):        'Kansas City',
    ('KS', 'johnson'):        'Kansas City',
    ('IN', 'marion'):         'Indianapolis',
    ('IN', 'hendricks'):      'Indianapolis',
    ('TN', 'davidson'):       'Nashville',
    ('TN', 'williamson'):     'Nashville',
    ('MN', 'hennepin'):       'Minneapolis',
    ('CO', 'denver'):         'Denver',
    ('CO', 'arapahoe'):       'Denver',
    ('MA', 'middlesex'):      'Boston',
    ('MA', 'suffolk'):        'Boston',
}


def _norm_county(s: str) -> str:
    """Lowercase, strip 'County' / 'Parish', collapse whitespace."""
    s = (s or '').strip().lower()
    s = re.sub(r'\b(county|parish|borough)\b', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def _norm_state(s: str) -> str:
    s = (s or '').strip().upper()
    return s[:2]


def map_to_metro(state: str, county: str) -> str:
    key = (_norm_state(state), _norm_county(county))
    return COUNTY_TO_METRO.get(key, 'Other')


# ── DC-related project name keywords ────────────────────────────────────
DC_KEYWORDS = (
    'data center', 'datacenter', 'data-center',
    'compute', 'cloud', 'colocation', 'colo',
    ' ai ', 'ai cluster', 'ai campus', 'gpu', 'hyperscale',
)
DC_TENANTS = (
    'microsoft', 'meta', 'facebook', 'google', 'alphabet', 'amazon', 'aws',
    'apple', 'oracle', 'openai', 'crusoe', 'xai', 'qts', 'equinix',
    'digital realty', 'cyrusone', 'aligned', 'compass', 'stack', 'vantage',
    'edgeconnex', 'switch', 'iron mountain', 'airtrunk', 'cloudhq',
    't5 data', 'sabey', 'prime data',
)


def is_dc_named(project_name: str) -> bool:
    n = (project_name or '').lower()
    return any(k in n for k in DC_KEYWORDS) or any(t in n for t in DC_TENANTS)


# ── PJM ──────────────────────────────────────────────────────────────────

def fetch_pjm() -> dict:
    """Fetch PJM's New Services Queue as a list of normalized rows."""
    try:
        r = requests.get(PJM_QUEUE_URL, timeout=30,
                         headers={'User-Agent': USER_AGENT, 'Accept': 'application/json'})
        if r.status_code != 200:
            return {'ok': False, 'error': f'HTTP {r.status_code}', 'iso': 'PJM'}
        data = r.json()
    except Exception as e:
        return {'ok': False, 'error': str(e), 'iso': 'PJM'}

    rows = []
    items = data if isinstance(data, list) else data.get('queue', data.get('items', []))
    for item in items:
        # PJM's keys vary slightly; defensively try common variants.
        name   = item.get('projectName') or item.get('name') or item.get('Project') or ''
        mw_raw = (item.get('mwCapacity') or item.get('mw') or item.get('MW')
                   or item.get('megawatts') or 0)
        try: mw = float(mw_raw)
        except (TypeError, ValueError): mw = 0.0
        state  = item.get('state') or item.get('State') or ''
        county = item.get('county') or item.get('County') or ''
        status = item.get('status') or item.get('Status') or ''
        ised   = item.get('inServiceDate') or item.get('expectedInServiceDate') or ''
        rows.append({
            'iso': 'PJM',
            'project_name': name,
            'mw': round(mw, 1),
            'state': _norm_state(state),
            'county': county,
            'metro': map_to_metro(state, county),
            'status': status,
            'in_service': ised,
            'is_dc_named': is_dc_named(name),
        })
    return {'ok': True, 'iso': 'PJM', 'count': len(rows), 'rows': rows}


# ── ERCOT ────────────────────────────────────────────────────────────────

def _discover_ercot_xlsx() -> str | None:
    try:
        r = requests.get(ERCOT_RESOURCE_PAGE, timeout=20,
                         headers={'User-Agent': USER_AGENT})
        if r.status_code != 200:
            return None
        m = ERCOT_XLSX_PATTERN.search(r.text)
        return m.group(0) if m else None
    except Exception:
        return None


def fetch_ercot() -> dict:
    """Discover the latest ERCOT generation interconnection XLSX and parse it."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {'ok': False, 'error': 'openpyxl not installed', 'iso': 'ERCOT'}

    url = _discover_ercot_xlsx()
    if not url:
        return {'ok': False, 'error': 'could not locate latest ERCOT XLSX', 'iso': 'ERCOT'}
    try:
        r = requests.get(url, timeout=60, headers={'User-Agent': USER_AGENT})
        if r.status_code != 200:
            return {'ok': False, 'error': f'HTTP {r.status_code}', 'iso': 'ERCOT', 'url': url}
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as e:
        return {'ok': False, 'error': str(e), 'iso': 'ERCOT', 'url': url}

    rows = []
    # ERCOT typically uses a single tab with columns like: INR, Project Name,
    # Interconnecting Entity, County, Capacity (MW), Status, COD ... we look
    # for the header row dynamically.
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c) if c is not None else '' for c in row]
            low = ' '.join(c.lower() for c in cells)
            if 'project name' in low and ('capacity' in low or 'mw' in low):
                header = cells
                continue
            if header is None:
                continue
            data = dict(zip(header, cells))
            # Defensive: try multiple field names.
            name   = (data.get('Project Name') or data.get('Project')
                       or data.get('Resource Name') or '').strip()
            cnty   = (data.get('County') or '').strip()
            mw_raw = (data.get('Capacity (MW)') or data.get('MW') or data.get('Capacity') or 0)
            try: mw = float(str(mw_raw).replace(',', '').strip() or 0)
            except ValueError: mw = 0.0
            status = (data.get('Status') or '').strip()
            ised   = (data.get('Approved for Energization') or data.get('COD')
                       or data.get('Projected COD') or '').strip()
            if not name and not cnty:
                continue
            rows.append({
                'iso': 'ERCOT',
                'project_name': name,
                'mw': round(mw, 1),
                'state': 'TX',
                'county': cnty,
                'metro': map_to_metro('TX', cnty),
                'status': status,
                'in_service': ised,
                'is_dc_named': is_dc_named(name),
            })
        if rows:
            break  # only need first matching sheet
    return {'ok': True, 'iso': 'ERCOT', 'count': len(rows), 'rows': rows, 'url': url}


# ── MISO ─────────────────────────────────────────────────────────────────

def fetch_miso() -> dict:
    """Fetch MISO's generator interconnection queue (CSV download endpoint)."""
    import csv as _csv
    try:
        r = requests.get(MISO_QUEUE_URL, timeout=60,
                         headers={'User-Agent': USER_AGENT, 'Accept': 'text/csv'})
        if r.status_code != 200:
            return {'ok': False, 'error': f'HTTP {r.status_code}', 'iso': 'MISO'}
        text = r.content.decode('utf-8-sig', errors='ignore')
    except Exception as e:
        return {'ok': False, 'error': str(e), 'iso': 'MISO'}

    rows = []
    reader = _csv.DictReader(io.StringIO(text))
    for d in reader:
        # MISO column names vary (queue export vs UI scrape). Accept common variants.
        name   = (d.get('Project Name') or d.get('Project') or d.get('Project Number') or '').strip()
        cnty   = (d.get('County 1') or d.get('County') or d.get('County (Site)') or '').strip()
        st     = (d.get('State 1') or d.get('State') or d.get('State (Site)') or '').strip()
        mw_raw = (d.get('Capacity (MW)') or d.get('Capacity') or d.get('Net MW')
                   or d.get('MW') or 0)
        try: mw = float(str(mw_raw).replace(',', '').strip() or 0)
        except ValueError: mw = 0.0
        status = (d.get('Status') or d.get('Queue Status') or '').strip()
        ised   = (d.get('Projected In-Service Date') or d.get('In Service Date')
                   or d.get('Projected COD') or '').strip()
        if not name and not cnty:
            continue
        rows.append({
            'iso': 'MISO',
            'project_name': name,
            'mw': round(mw, 1),
            'state': _norm_state(st),
            'county': cnty,
            'metro': map_to_metro(st, cnty),
            'status': status,
            'in_service': ised,
            'is_dc_named': is_dc_named(name),
        })
    return {'ok': True, 'iso': 'MISO', 'count': len(rows), 'rows': rows}


# ── CAISO ────────────────────────────────────────────────────────────────

def _discover_caiso_xlsx() -> str | None:
    try:
        r = requests.get(CAISO_RESOURCE_PAGE, timeout=20,
                         headers={'User-Agent': USER_AGENT})
        if r.status_code != 200:
            return None
        m = CAISO_XLSX_PATTERN.search(r.text)
        return m.group(0) if m else None
    except Exception:
        return None


def fetch_caiso() -> dict:
    """Discover and parse the latest CAISO Public Queue Report XLSX."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {'ok': False, 'error': 'openpyxl not installed', 'iso': 'CAISO'}

    url = _discover_caiso_xlsx()
    if not url:
        return {'ok': False, 'error': 'could not locate latest CAISO XLSX', 'iso': 'CAISO'}
    try:
        r = requests.get(url, timeout=60, headers={'User-Agent': USER_AGENT})
        if r.status_code != 200:
            return {'ok': False, 'error': f'HTTP {r.status_code}', 'iso': 'CAISO', 'url': url}
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as e:
        return {'ok': False, 'error': str(e), 'iso': 'CAISO', 'url': url}

    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else '' for c in row]
            low = ' '.join(c.lower() for c in cells)
            if 'project name' in low and ('county' in low or 'mw' in low or 'capacity' in low):
                header = cells
                continue
            if header is None:
                continue
            d = dict(zip(header, cells))
            name   = (d.get('Project Name') or d.get('Project') or '').strip()
            cnty   = (d.get('County') or d.get('Site County') or '').strip()
            st     = 'CA'  # CAISO is California
            mw_raw = (d.get('MW') or d.get('Capacity (MW)') or d.get('Net MW') or 0)
            try: mw = float(str(mw_raw).replace(',', '').strip() or 0)
            except ValueError: mw = 0.0
            status = (d.get('Status') or d.get('Queue Status') or '').strip()
            ised   = (d.get('In-Service Date') or d.get('Commercial Operation Date')
                       or d.get('Projected COD') or '').strip()
            if not name and not cnty:
                continue
            rows.append({
                'iso': 'CAISO',
                'project_name': name,
                'mw': round(mw, 1),
                'state': st,
                'county': cnty,
                'metro': map_to_metro(st, cnty),
                'status': status,
                'in_service': ised,
                'is_dc_named': is_dc_named(name),
            })
        if rows:
            break
    return {'ok': True, 'iso': 'CAISO', 'count': len(rows), 'rows': rows, 'url': url}


# ── Aggregation ──────────────────────────────────────────────────────────

def summarize_by_metro(rows: list[dict]) -> list[dict]:
    """Group rows by metro and return per-metro totals."""
    by_metro: dict[str, dict] = {}
    for r in rows:
        m = r['metro']
        b = by_metro.setdefault(m, {
            'metro': m,
            'queue_mw_total': 0.0,
            'dc_named_mw': 0.0,
            'project_count': 0,
            'dc_named_count': 0,
            'top_projects': [],
        })
        b['queue_mw_total'] += r['mw']
        b['project_count']  += 1
        if r['is_dc_named']:
            b['dc_named_mw']    += r['mw']
            b['dc_named_count'] += 1
            b['top_projects'].append({
                'name': r['project_name'],
                'mw': r['mw'],
                'county': r['county'],
                'in_service': r['in_service'],
                'iso': r['iso'],
            })
    for b in by_metro.values():
        b['queue_mw_total'] = round(b['queue_mw_total'], 1)
        b['dc_named_mw']    = round(b['dc_named_mw'], 1)
        # Top 10 DC-named projects per metro by MW
        b['top_projects'] = sorted(b['top_projects'], key=lambda x: x['mw'], reverse=True)[:10]
    return sorted(by_metro.values(), key=lambda x: -x['queue_mw_total'])


def pull_all() -> dict:
    """Fetch PJM + ERCOT + MISO + CAISO, combine, summarize."""
    pjm   = fetch_pjm()
    ercot = fetch_ercot()
    miso  = fetch_miso()
    caiso = fetch_caiso()
    rows = []
    for r in (pjm, ercot, miso, caiso):
        if r.get('ok'):
            rows.extend(r['rows'])
    summary = summarize_by_metro(rows)
    out = {
        'ok': any(r.get('ok') for r in (pjm, ercot, miso, caiso)),
        'pjm':   {k: v for k, v in pjm.items()   if k != 'rows'},
        'ercot': {k: v for k, v in ercot.items() if k != 'rows'},
        'miso':  {k: v for k, v in miso.items()  if k != 'rows'},
        'caiso': {k: v for k, v in caiso.items() if k != 'rows'},
        'total_rows': len(rows),
        'by_metro': summary,
    }
    # Record freshness on success.
    try:
        from backend.data_centers import freshness
        freshness.record_pull('iso_queues', {
            'isos_ok': [k for k, v in {'pjm': pjm, 'ercot': ercot, 'miso': miso, 'caiso': caiso}.items() if v.get('ok')],
            'rows': len(rows),
        })
    except Exception:
        pass
    return out


def to_csv(rows_or_summary) -> str:
    """Flatten by-metro summary into CSV the admin can download.
    Accepts either the raw rows or the summarized list."""
    import csv as _csv
    import io as _io
    buf = _io.StringIO()
    w = _csv.writer(buf)
    # Summary mode
    if rows_or_summary and isinstance(rows_or_summary[0], dict) and 'queue_mw_total' in rows_or_summary[0]:
        w.writerow(['metro', 'queue_mw_total', 'dc_named_mw', 'project_count',
                    'dc_named_count', 'top_dc_projects'])
        for b in rows_or_summary:
            top = '; '.join(f"{p['name']} ({p['mw']} MW, {p['county']})" for p in b['top_projects'])
            w.writerow([b['metro'], b['queue_mw_total'], b['dc_named_mw'],
                        b['project_count'], b['dc_named_count'], top])
    else:
        w.writerow(['iso', 'project_name', 'mw', 'state', 'county', 'metro',
                    'status', 'in_service', 'is_dc_named'])
        for r in rows_or_summary:
            w.writerow([r['iso'], r['project_name'], r['mw'], r['state'], r['county'],
                        r['metro'], r['status'], r['in_service'], 'yes' if r['is_dc_named'] else ''])
    return buf.getvalue()
