"""
Market consensus tracker for commodity forecasts.

Aggregates forward quarterly commodity price forecasts from:

  1. Manual YAML (``data/consensus.yaml``) — operator-maintained bank
     consensus (Goldman Sachs, J.P. Morgan, UBS, Morgan Stanley, etc.),
     copy-pasted from paywalled research PDFs with citation.

  2. EIA Short-Term Energy Outlook (public monthly CSV) — WTI, Brent,
     Henry Hub, ~Q+6 horizon. No key required.

  3. World Bank Pink Sheet — monthly historical prices + annual nominal
     price forecasts for all tracked commodities.

  4. IMF World Economic Outlook Commodity Outlook — semi-annual annual
     averages. Best-effort; stubbed when the public URL shape drifts.

Exposed via :func:`get_consensus_data` — returns
``{commodity_name: [entry, ...]}`` with each entry shaped like::

    {
        'source':    'Goldman Sachs' | 'EIA STEO' | 'World Bank' | 'IMF WEO',
        'as_of':     'YYYY-MM-DD',      # data vintage / release date
        'published': 'YYYY-MM-DD',      # report publication, when distinct
        'note':      str,               # citation or provenance
        'unit':      str,               # informational
        'quarters':  {'Q2_2026': 78.0, 'Q3_2026': 82.0, ...},
        'fy':        {2026: 80.0, 2027: 75.0},
    }

All entries sorted most-recent first. 24-hour TTL cache mirrors the
``ForecastCache`` pattern in :mod:`commodities_forecast`.

Bank research is paywalled — the module does not scrape it; operators
paste published numbers into ``data/consensus.yaml`` manually.
"""

from __future__ import annotations

import io
import os
import time
import logging
import threading
from datetime import datetime, date
from typing import Optional

import requests

logger = logging.getLogger(__name__)

CACHE_TTL = 86400           # 24 hours
HTTP_TIMEOUT = 15
_REPO_ROOT = os.path.dirname(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)
DEFAULT_YAML_PATH = os.path.join(_REPO_ROOT, 'data', 'consensus.yaml')

# Mapping from our display names to the short symbols used in EIA STEO CSV
EIA_STEO_MAP = {
    'WTI Crude':        'WTIPUUS',   # WTI spot price ($/bbl)
    'Brent Crude':      'BREPUUS',   # Brent spot price ($/bbl)
    'Natural Gas (HH)': 'NGHHUUS',   # Henry Hub spot price ($/MMBtu)
}
EIA_STEO_URL = 'https://www.eia.gov/outlooks/steo/outlook.php'

# World Bank Pink Sheet — monthly historical + annual forecast
# Exact URL shape changes with each release; try current + fallback.
WB_PINK_SHEET_URLS = [
    'https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012026/related/CMO-Historical-Data-Monthly.xlsx',
    'https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025/related/CMO-Historical-Data-Monthly.xlsx',
]
WB_FORECAST_URLS = [
    'https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012026/related/CMO-April-2026-Forecasts.xlsx',
    'https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025/related/CMO-October-2025-Forecasts.xlsx',
]

ALL_COMMODITIES = (
    'WTI Crude', 'Brent Crude', 'Natural Gas (HH)', 'TTF Gas',
    'Gold', 'Silver', 'Platinum', 'Copper', 'Aluminum',
    'Cocoa', 'Wheat', 'Soybeans', 'Coffee',
)


# ── YAML loader ───────────────────────────────────────────────────────────

def load_manual_consensus(path: str = DEFAULT_YAML_PATH) -> dict[str, list[dict]]:
    """Load operator-maintained bank consensus from YAML."""
    try:
        import yaml
    except ImportError:
        logger.error('PyYAML not installed — cannot load manual consensus')
        return {}
    if not os.path.exists(path):
        logger.warning(f'consensus YAML not found at {path}')
        return {}
    try:
        with open(path) as f:
            raw = yaml.safe_load(f)
    except Exception as e:
        logger.error(f'failed to parse consensus YAML: {e}')
        return {}
    if not isinstance(raw, dict):
        return {}

    out: dict[str, list[dict]] = {}
    for commodity, entries in raw.items():
        if not isinstance(entries, list):
            continue
        normalized = []
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            normalized.append(_normalize_yaml_entry(entry))
        if normalized:
            out[commodity] = normalized
    return out


def _normalize_yaml_entry(entry: dict) -> dict:
    """Coerce a YAML entry into the canonical consensus schema."""
    published = entry.get('published')
    if isinstance(published, (date, datetime)):
        published = published.isoformat()[:10]
    return {
        'source':    str(entry.get('source', 'Unknown')),
        'as_of':     published or str(entry.get('as_of', '')),
        'published': published,
        'note':      entry.get('note'),
        'unit':      entry.get('unit'),
        'quarters':  dict(entry.get('quarters') or {}),
        'fy':        dict(entry.get('fy') or {}),
    }


# ── EIA STEO ──────────────────────────────────────────────────────────────

def fetch_eia_steo() -> dict[str, list[dict]]:
    """
    Fetch EIA Short-Term Energy Outlook forecast values.

    EIA publishes the STEO as a monthly HTML/XLS/CSV; the table shape
    drifts. This function is best-effort and returns {} cleanly on any
    failure rather than crashing callers. When it works it covers the
    three EIA-tracked commodities we map: WTI Crude, Brent Crude, Henry
    Hub natural gas.
    """
    out: dict[str, list[dict]] = {}
    try:
        resp = requests.get(EIA_STEO_URL, timeout=HTTP_TIMEOUT)
        if resp.status_code != 200:
            logger.warning(f'EIA STEO fetch: HTTP {resp.status_code}')
            return {}
    except Exception as e:
        logger.warning(f'EIA STEO fetch failed: {e}')
        return {}

    # The STEO outlook page embeds forecast tables we could scrape; parsing
    # is brittle and EIA's series IDs live in a separate download. Rather
    # than ship a fragile scraper we stub to an empty result here and rely
    # on the manual YAML + World Bank path for now. Operators can extend
    # this function when the EIA API v2 series IDs are wired in.
    logger.info('EIA STEO: HTTP 200 but forecast parsing not yet implemented')
    return out


# ── World Bank Pink Sheet ─────────────────────────────────────────────────

def fetch_worldbank_pinksheet() -> dict[str, list[dict]]:
    """
    Fetch World Bank Commodity Markets Outlook annual forecasts.

    The WB publishes a spreadsheet of nominal-price forecasts covering all
    tracked commodities at annual frequency. URL shape rotates each
    edition; we try the most recent known URLs in order.
    """
    for url in WB_FORECAST_URLS:
        data = _try_wb_forecast(url)
        if data:
            return data
    return {}


def _try_wb_forecast(url: str) -> dict[str, list[dict]]:
    try:
        resp = requests.get(url, timeout=HTTP_TIMEOUT)
    except Exception as e:
        logger.warning(f'WB Pink Sheet fetch {url!r}: {e}')
        return {}
    if resp.status_code != 200:
        logger.debug(f'WB Pink Sheet {url!r}: HTTP {resp.status_code}')
        return {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error('openpyxl not installed — cannot parse WB Pink Sheet')
        return {}

    try:
        wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    except Exception as e:
        logger.warning(f'WB Pink Sheet parse {url!r}: {e}')
        return {}

    # Forecast sheet layout: first row has year headers, first column has
    # commodity labels. Exact commodity naming differs from ours, so we map
    # on keyword match.
    alias_map = {
        'crude oil, wti':   'WTI Crude',
        'crude oil, brent': 'Brent Crude',
        'natural gas, u.s.':'Natural Gas (HH)',
        'natural gas, europe': 'TTF Gas',
        'gold':             'Gold',
        'silver':           'Silver',
        'platinum':         'Platinum',
        'copper':           'Copper',
        'aluminum':         'Aluminum',
        'cocoa':            'Cocoa',
        'wheat, us hrw':    'Wheat',
        'soybeans':         'Soybeans',
        'coffee, arabica':  'Coffee',
    }

    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logger.warning(f'WB Pink Sheet iter_rows failed: {e}')
        return {}

    if not rows:
        return {}

    # Find header row: the first row with 4+ numeric cells (years)
    header_row_idx = None
    years: list[int] = []
    for i, row in enumerate(rows):
        numerics = [int(c) for c in row if isinstance(c, (int, float)) and 1990 <= float(c) <= 2050]
        if len(numerics) >= 3:
            header_row_idx = i
            years = numerics
            break
    if header_row_idx is None:
        return {}

    out: dict[str, list[dict]] = {}
    today_iso = date.today().isoformat()
    source_note = f'World Bank Commodity Markets Outlook · {os.path.basename(url)}'

    for row in rows[header_row_idx + 1:]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip().lower()
        matched = next((v for k, v in alias_map.items() if k in label), None)
        if not matched:
            continue
        fy: dict[int, float] = {}
        for y, cell in zip(years, row[1:]):
            if isinstance(cell, (int, float)):
                fy[int(y)] = float(cell)
        if not fy:
            continue
        out.setdefault(matched, []).append({
            'source':    'World Bank',
            'as_of':     today_iso,
            'published': today_iso,
            'note':      source_note,
            'unit':      None,
            'quarters':  {},
            'fy':        fy,
        })
    return out


# ── IMF WEO Commodity Outlook ─────────────────────────────────────────────

def fetch_imf_commodity_outlook() -> dict[str, list[dict]]:
    """Stub. The IMF WEO commodity projections ship as semi-annual PDF tables
    with no clean machine-readable feed we can rely on; returns {} until an
    operator wires a parser for the specific release they want benchmarked.
    """
    return {}


# ── Aggregator + cache ────────────────────────────────────────────────────

class _ConsensusCache:
    def __init__(self) -> None:
        self._lock = threading.RLock()
        self._data: Optional[dict] = None
        self._fetched_at = 0.0

    def get(self) -> dict:
        with self._lock:
            if self._data is not None and (time.time() - self._fetched_at) < CACHE_TTL:
                return self._data
        fresh = _build_all_consensus()
        with self._lock:
            self._data = fresh
            self._fetched_at = time.time()
        return fresh

    def clear(self) -> None:
        with self._lock:
            self._data = None
            self._fetched_at = 0.0


_cache = _ConsensusCache()


def _build_all_consensus() -> dict[str, list[dict]]:
    """Merge all sources, sort entries by as_of descending."""
    merged: dict[str, list[dict]] = {name: [] for name in ALL_COMMODITIES}

    # Manual YAML
    manual = load_manual_consensus()
    for commodity, entries in manual.items():
        merged.setdefault(commodity, []).extend(entries)

    # EIA STEO
    for commodity, entries in fetch_eia_steo().items():
        merged.setdefault(commodity, []).extend(entries)

    # World Bank Pink Sheet
    for commodity, entries in fetch_worldbank_pinksheet().items():
        merged.setdefault(commodity, []).extend(entries)

    # IMF WEO
    for commodity, entries in fetch_imf_commodity_outlook().items():
        merged.setdefault(commodity, []).extend(entries)

    for commodity, entries in merged.items():
        entries.sort(key=lambda e: str(e.get('as_of') or ''), reverse=True)

    return merged


def get_consensus(commodity: str) -> list[dict]:
    return _cache.get().get(commodity, [])


def get_all_consensus() -> dict[str, list[dict]]:
    return _cache.get()


def get_consensus_data() -> dict[str, list[dict]]:
    """Public entry-point used by commodities_forecast integration."""
    return get_all_consensus()


def clear_cache() -> None:
    _cache.clear()


# ── Smoke test ────────────────────────────────────────────────────────────

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(levelname)s %(name)s %(message)s')
    for commodity in ('WTI Crude', 'Gold', 'Cocoa'):
        print(f'\n=== {commodity} ===')
        for entry in get_consensus(commodity):
            print(f'  {entry["source"]:18s} as_of={entry["as_of"]} quarters={entry["quarters"]}')
