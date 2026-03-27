"""
Geopolitical Risk (GPR) Index by Iacoviello (2022).

Monthly country-level index for 44 countries, published by the Federal Reserve.
The academic standard benchmark for geopolitical risk measurement.
Downloaded from https://www.matteoiacoviello.com/gpr.htm

Used as a calibration signal within the base score calculation
alongside WGI governance scores and macro indicators.

Cached locally for 30 days. No API key needed — public data.
"""

import os
import json
import math
import logging
import requests
from datetime import datetime
from config import Config

logger = logging.getLogger(__name__)

# Monthly country-level data (Excel format)
GPR_DATA_URL = 'https://www.matteoiacoviello.com/gpr_files/data_gpr_export.xls'

GPR_CACHE_FILE = os.path.join(Config.DATA_DIR, 'gpr_cache.json')
GPR_CACHE_DAYS = 30

# GPR column names → ISO alpha-2 codes (44 countries)
GPR_COUNTRIES = {
    'GPRC_ARG': 'AR', 'GPRC_AUS': 'AU', 'GPRC_BRA': 'BR',
    'GPRC_CAN': 'CA', 'GPRC_CHN': 'CN', 'GPRC_COL': 'CO',
    'GPRC_EGY': 'EG', 'GPRC_FRA': 'FR', 'GPRC_DEU': 'DE',
    'GPRC_IND': 'IN', 'GPRC_IDN': 'ID', 'GPRC_ISR': 'IL',
    'GPRC_ITA': 'IT', 'GPRC_JPN': 'JP', 'GPRC_KOR': 'KR',
    'GPRC_MEX': 'MX', 'GPRC_MYS': 'MY', 'GPRC_NGA': 'NG',
    'GPRC_PAK': 'PK', 'GPRC_PHL': 'PH', 'GPRC_POL': 'PL',
    'GPRC_RUS': 'RU', 'GPRC_SAU': 'SA', 'GPRC_ZAF': 'ZA',
    'GPRC_ESP': 'ES', 'GPRC_SWE': 'SE', 'GPRC_THA': 'TH',
    'GPRC_TUR': 'TR', 'GPRC_GBR': 'GB', 'GPRC_USA': 'US',
    'GPRC_UKR': 'UA', 'GPRC_VEN': 'VE', 'GPRC_IRN': 'IR',
    'GPRC_IRQ': 'IQ', 'GPRC_NOR': 'NO', 'GPRC_NLD': 'NL',
    'GPRC_CHE': 'CH', 'GPRC_CHL': 'CL', 'GPRC_PER': 'PE',
    'GPRC_COD': 'CD', 'GPRC_TUN': 'TN', 'GPRC_GRC': 'GR',
    'GPRC_BGD': 'BD', 'GPRC_HUN': 'HU',
}

# In-memory cache
_gpr_scores = None
_gpr_fetched_at = None


def _load_cache():
    """Load GPR scores from disk cache if fresh enough."""
    if not os.path.exists(GPR_CACHE_FILE):
        return None
    try:
        with open(GPR_CACHE_FILE, 'r') as f:
            cache = json.load(f)
        fetched = datetime.fromisoformat(cache.get('fetched_at', '2000-01-01'))
        if (datetime.utcnow() - fetched).days < GPR_CACHE_DAYS:
            return cache.get('scores', {})
    except Exception:
        pass
    return None


def _save_cache(scores):
    """Save GPR scores to disk cache."""
    try:
        os.makedirs(Config.DATA_DIR, exist_ok=True)
        cache = {
            'fetched_at': datetime.utcnow().isoformat(),
            'scores': scores,
        }
        with open(GPR_CACHE_FILE, 'w') as f:
            json.dump(cache, f, separators=(',', ':'))
    except Exception as e:
        logger.warning(f"Failed to save GPR cache: {e}")


def _normalize_gpr(raw_value):
    """
    Normalize raw GPR index value to 0-100 scale.
    GPR typically ranges from ~50 (low risk) to ~500+ (crisis).
    Uses log2 scaling for better spread:
      GPR  50 → ~10
      GPR 100 → ~27
      GPR 200 → ~44
      GPR 400 → ~61
      GPR 800 → ~78
    """
    if raw_value is None or raw_value <= 0:
        return 0.0
    log_val = math.log2(max(raw_value, 1))
    # Map log2(50)≈5.64 to ~10, log2(800)≈9.64 to ~78
    norm = max(0, min(100, (log_val - 5.0) * 17.0))
    return round(norm, 1)


def fetch_gpr_data():
    """
    Download and parse GPR index data. Returns {iso2: normalized_score}.
    Uses disk cache (30 days), then memory cache.
    """
    global _gpr_scores, _gpr_fetched_at

    # Check memory cache
    if _gpr_scores is not None and _gpr_fetched_at is not None:
        if (datetime.utcnow() - _gpr_fetched_at).days < GPR_CACHE_DAYS:
            return _gpr_scores

    # Check disk cache
    cached = _load_cache()
    if cached:
        _gpr_scores = cached
        _gpr_fetched_at = datetime.utcnow()
        return cached

    # Download fresh data
    try:
        resp = requests.get(GPR_DATA_URL, timeout=30)
        if resp.status_code != 200:
            logger.warning(f"GPR download failed: HTTP {resp.status_code}")
            return _gpr_scores or {}

        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
        ws = wb.active

        # Parse headers from first row
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            headers = [cell.value for cell in row]
            break

        if not headers:
            logger.warning("GPR Excel has no headers")
            return _gpr_scores or {}

        # Find the last data row (most recent month)
        last_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            last_row = row

        wb.close()

        if not last_row:
            logger.warning("GPR Excel has no data rows")
            return _gpr_scores or {}

        # Extract and normalize scores
        scores = {}
        for i, header in enumerate(headers):
            if header and header in GPR_COUNTRIES and i < len(last_row):
                raw = last_row[i]
                if raw is not None:
                    try:
                        iso2 = GPR_COUNTRIES[header]
                        scores[iso2] = _normalize_gpr(float(raw))
                    except (ValueError, TypeError):
                        pass

        if scores:
            _gpr_scores = scores
            _gpr_fetched_at = datetime.utcnow()
            _save_cache(scores)
            logger.info(f"GPR Index loaded: {len(scores)} countries")
        else:
            logger.warning("GPR parsing produced no scores")

        return scores

    except ImportError:
        logger.warning("openpyxl not installed — GPR Index unavailable")
        return _gpr_scores or {}
    except Exception as e:
        logger.error(f"Failed to fetch GPR data: {e}")
        return _gpr_scores or {}


def get_gpr_score(country_code):
    """
    Get normalized GPR score (0-100) for a country.
    Returns None if country not covered by GPR (44 countries only).
    """
    scores = fetch_gpr_data()
    return scores.get(country_code.upper())
