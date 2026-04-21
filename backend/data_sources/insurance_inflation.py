"""
Insurance/Reinsurance Inflation data source.

Pulls ~35 time series from public ONS and Eurostat APIs, organized into
6 insurance line categories: Medical, Legal, Insurance, Bodily Injury,
Fire & Allied, Auto Physical Damage.

Data sources:
  ONS: CPIH (mm23), AWE (emp), PPI (mm22), SPPI (sppi), Construction OPI (Excel)
  Eurostat: HICP (prc_hicp_manr), PPI (sts_inpp_m, sts_inppd_m),
            Construction (sts_copi_q), Labour Costs (lc_lci_r2_q)

All values returned as YoY % change to match Haver output format.
Thread-safe cache with 24-hour TTL.
"""

import io
import threading
import time
import logging
import requests
from datetime import datetime

logger = logging.getLogger(__name__)

CACHE_TTL = 86400   # 24 hours
RETRY_BACKOFF = 3600  # 1 hour after failure
USER_AGENT = 'Mozilla/5.0 (compatible; ParraMacro/1.0)'

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Series definitions
# ══════════════════════════════════════════════════════════════════════════════

# ── ONS URL patterns ─────────────────────────────────────────────────────────
ONS_URLS = {
    'mm23': 'https://www.ons.gov.uk/economy/inflationandpriceindices/timeseries/{cdid}/mm23/data',
    'emp':  'https://www.ons.gov.uk/employmentandlabourmarket/peopleinwork/earningsandworkinghours/timeseries/{cdid}/emp/data',
    'mm22': 'https://www.ons.gov.uk/economy/inflationandpriceindices/timeseries/{cdid}/mm22/data',
    'sppi': 'https://www.ons.gov.uk/economy/inflationandpriceindices/timeseries/{cdid}/sppi/data',
}

# ONS series: key → {cdid, dataset, label, color, category}
# All return index values — YoY computed in backend
ONS_SERIES = {
    # Medical
    'uk_medical_svc':    {'cdid': 'L53Y', 'dataset': 'mm23', 'label': 'UK CPIH: Medical & Paramedical Services', 'color': '#ef4444', 'category': 'medical'},
    'uk_hospital':       {'cdid': 'L542', 'dataset': 'mm23', 'label': 'UK CPIH: Hospital Services',              'color': '#f87171', 'category': 'medical'},
    'uk_pharma':         {'cdid': 'L53V', 'dataset': 'mm23', 'label': 'UK CPIH: Pharmaceutical Products',        'color': '#fca5a5', 'category': 'medical'},
    'uk_med_equip':      {'cdid': 'L53W', 'dataset': 'mm23', 'label': 'UK CPIH: Other Medical Equipment',        'color': '#fb923c', 'category': 'medical'},
    'uk_dental':         {'cdid': 'L53Z', 'dataset': 'mm23', 'label': 'UK CPIH: Dental Services',                'color': '#fdba74', 'category': 'medical'},
    # Legal
    'uk_sppi_legal':     {'cdid': 'HSGL', 'dataset': 'sppi', 'label': 'UK SPPI: Legal Services',                 'color': '#f59e0b', 'category': 'legal'},
    'uk_awe_prof':       {'cdid': 'K5EC', 'dataset': 'emp',  'label': 'UK AWE: Prof, Scientific & Technical (legal proxy)', 'color': '#eab308', 'category': 'legal', 'approximate': True},
    # Insurance
    'uk_awe_finance':    {'cdid': 'K58I', 'dataset': 'emp',  'label': 'UK AWE: Finance & Insurance',             'color': '#a855f7', 'category': 'insurance'},
    # Bodily Injury
    'uk_awe_total':      {'cdid': 'KAB9', 'dataset': 'emp',  'label': 'UK AWE: Whole Economy Total Pay',         'color': '#3b82f6', 'category': 'bodily_injury'},
    # Fire & Allied
    'uk_maint_repair':   {'cdid': 'L537', 'dataset': 'mm23', 'label': 'UK CPIH: Maintenance & Repair of Dwelling','color': '#10b981', 'category': 'fire_allied'},
    'uk_ppi_cement':     {'cdid': 'GHGF', 'dataset': 'mm22', 'label': 'UK PPI: Cement, Lime & Plaster',          'color': '#34d399', 'category': 'fire_allied'},
    'uk_ppi_glass':      {'cdid': 'GHGH', 'dataset': 'mm22', 'label': 'UK PPI: Glass & Clay Products',           'color': '#6ee7b7', 'category': 'fire_allied'},
    'uk_awe_construction': {'cdid': 'K583', 'dataset': 'emp', 'label': 'UK AWE: Construction',                   'color': '#059669', 'category': 'fire_allied'},
    # Auto Physical Damage
    'uk_vehicle_maint':  {'cdid': 'L54A', 'dataset': 'mm23', 'label': 'UK CPIH: Vehicle Maintenance & Repairs',  'color': '#06b6d4', 'category': 'auto_physical'},
    'uk_vehicle_parts':  {'cdid': 'L548', 'dataset': 'mm23', 'label': 'UK CPIH: Spare Parts & Accessories',      'color': '#22d3ee', 'category': 'auto_physical'},
    'uk_awe_transport':  {'cdid': 'K58F', 'dataset': 'emp',  'label': 'UK AWE: Transport & Storage',             'color': '#67e8f9', 'category': 'auto_physical'},
}

# ── Eurostat datasets ────────────────────────────────────────────────────────
EUROSTAT_BASE = 'https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data'

# HICP — annual rate of change (no YoY computation needed)
EU_HICP_SERIES = {
    'eu_medical_svc':   {'coicop': 'CP0621', 'label': 'EU HICP: Medical & Paramedical Services', 'color': '#ef4444', 'category': 'medical'},
    'eu_hospital':      {'coicop': 'CP063',  'label': 'EU HICP: Hospital Services',              'color': '#f87171', 'category': 'medical'},
    'eu_pharma':        {'coicop': 'CP0611', 'label': 'EU HICP: Pharmaceutical Products',        'color': '#fca5a5', 'category': 'medical'},
    'eu_med_equip':     {'coicop': 'CP0613', 'label': 'EU HICP: Therapeutic Appliances & Equip', 'color': '#fb923c', 'category': 'medical'},
    'eu_dental':        {'coicop': 'CP0622', 'label': 'EU HICP: Dental Services',                'color': '#fdba74', 'category': 'medical'},
    'eu_insurance_svc': {'coicop': 'CP125',  'label': 'EU HICP: Insurance Services',              'color': '#a855f7', 'category': 'insurance'},
    'eu_financial_svc': {'coicop': 'CP126',  'label': 'EU HICP: Financial Services n.e.c.',       'color': '#9333ea', 'category': 'insurance'},
    'eu_maint_repair':  {'coicop': 'CP043',  'label': 'EU HICP: Maintenance & Repair of Dwelling','color': '#10b981', 'category': 'fire_allied'},
    'eu_vehicle_maint': {'coicop': 'CP0723', 'label': 'EU HICP: Vehicle Maintenance & Repairs',  'color': '#06b6d4', 'category': 'auto_physical'},
    'eu_vehicle_parts': {'coicop': 'CP0721', 'label': 'EU HICP: Spare Parts & Accessories',      'color': '#22d3ee', 'category': 'auto_physical'},
}

# Country-specific legal proxies (approximate — CP127 "Other Services n.e.c."
# is the closest HICP category to legal services costs)
EU_LEGAL_NL = {'nl_legal': {'coicop': 'CP127', 'label': 'NL HICP: Other Services (legal proxy)', 'color': '#eab308', 'category': 'legal', 'approximate': True}}
EU_LEGAL_IT = {'it_legal': {'coicop': 'CP127', 'label': 'IT HICP: Other Services (legal proxy)',  'color': '#ca8a04', 'category': 'legal', 'approximate': True}}
EU_LEGAL_DE = {'de_legal': {'coicop': 'CP127', 'label': 'DE HICP: Other Services (legal proxy)',  'color': '#f59e0b', 'category': 'legal', 'approximate': True}}
EU_LEGAL_FR = {'fr_legal': {'coicop': 'CP127', 'label': 'FR HICP: Other Services (legal proxy)',  'color': '#d97706', 'category': 'legal', 'approximate': True}}

# PPI — use annual rate of change (unit=PCH_PRE for YoY)
EU_PPI_SERIES = {
    'eu_ppi_cement': {'nace': 'C235', 'dataset': 'sts_inpp_m',  'label': 'EU PPI: Cement, Lime & Plaster',  'color': '#34d399', 'category': 'fire_allied'},
    'eu_ppi_glass':  {'nace': 'C231', 'dataset': 'sts_inppd_m', 'label': 'EU PPI Domestic: Glass Products', 'color': '#6ee7b7', 'category': 'fire_allied'},
}

# Construction — quarterly (unit=PCH_Q4 = YoY quarterly change)
# Dataset sts_copi_q covers CPA_F41001_X_410014: Residential buildings excl residences for communities
EU_CONSTRUCTION = {
    'eu_cci_res':   {'indic': 'COST',    'label': 'EU CCI: Residential Buildings (excl Communities)',   'color': '#f97316', 'category': 'fire_allied'},
    'eu_cppi_res':  {'indic': 'PRC_PRR', 'label': 'EU CPPI: Residential Buildings (excl Communities)', 'color': '#fb923c', 'category': 'fire_allied'},
}

# Labour Cost Index — quarterly
EU_LCI = {
    'eu_lci_finance':       {'nace': 'K',   'label': 'EU LCI: Finance & Insurance',           'color': '#8b5cf6', 'category': 'insurance'},
    'eu_lci_total':         {'nace': 'B-S', 'label': 'EU LCI: Total Business Economy',        'color': '#a78bfa', 'category': 'bodily_injury'},
    'eu_lci_construction':  {'nace': 'F',   'label': 'EU LCI: Construction',                   'color': '#c084fc', 'category': 'fire_allied'},
    'eu_lci_transport':     {'nace': 'H',   'label': 'EU LCI: Transport & Storage',            'color': '#d8b4fe', 'category': 'auto_physical'},
    'eu_lci_prof_services': {'nace': 'M',   'label': 'EU LCI: Professional & Scientific Svcs', 'color': '#fbbf24', 'category': 'legal'},
    'eu_lci_industry':      {'nace': 'B-E', 'label': 'EU LCI: Industry',                      'color': '#64748b', 'category': 'bodily_injury'},
    'eu_lci_construction_bi': {'nace': 'F', 'label': 'EU LCI: Construction',                   'color': '#94a3b8', 'category': 'bodily_injury'},
    'eu_lci_services':      {'nace': 'G-N', 'label': 'EU LCI: Services',                      'color': '#cbd5e1', 'category': 'bodily_injury'},
}

# ── US BLS series ────────────────────────────────────────────────────────────
# CPI-U, PPI, and CES/AWE series from Bureau of Labor Statistics API v2
# All return index values — YoY computed in backend via _compute_yoy()
BLS_API_URL = 'https://api.bls.gov/publicAPI/v2/timeseries/data/'

BLS_SERIES = {
    # Medical (CPI-U)
    'us_physicians':     {'bls_id': 'CUUR0000SEMC01', 'label': 'US CPI: Physicians\' Services',         'color': '#dc2626', 'category': 'medical'},
    'us_hospital':       {'bls_id': 'CUUR0000SEMD01', 'label': 'US CPI: Hospital Services',              'color': '#b91c1c', 'category': 'medical'},
    'us_drugs':          {'bls_id': 'CUUR0000SEMF',   'label': 'US CPI: Medicinal Drugs',                'color': '#991b1b', 'category': 'medical'},
    'us_med_equip':      {'bls_id': 'CUUR0000SEMG',   'label': 'US CPI: Medical Equipment & Supplies',   'color': '#7f1d1d', 'category': 'medical'},
    'us_dental':         {'bls_id': 'CUUR0000SEMC02', 'label': 'US CPI: Dental Services',                'color': '#450a0a', 'category': 'medical'},
    'us_nursing':        {'bls_id': 'CUUR0000SEMD02', 'label': 'US CPI: Nursing Homes & Adult Daycare',  'color': '#fee2e2', 'category': 'medical'},
    'us_other_med':      {'bls_id': 'CUUR0000SEMC04', 'label': 'US CPI: Other Medical Professionals',    'color': '#fecaca', 'category': 'medical'},
    # Fire & Allied (PPI + AWE)
    'us_ppi_res_constr': {'bls_id': 'WPUIP2311001',   'label': 'US PPI: Inputs to Residential Construction', 'color': '#065f46', 'category': 'fire_allied'},
    'us_ppi_cement':     {'bls_id': 'WPU1322',        'label': 'US PPI: Cement, Hydraulic',              'color': '#047857', 'category': 'fire_allied'},
    'us_ppi_glass':      {'bls_id': 'WPU1311',        'label': 'US PPI: Flat Glass',                     'color': '#059669', 'category': 'fire_allied'},
    'us_ppi_construction': {'bls_id': 'WPUFD43',      'label': 'US PPI: Final Demand Construction',      'color': '#0d9488', 'category': 'fire_allied'},
    'us_ppi_bldg_maint': {'bls_id': 'PCU2381MR2381MR','label': 'US PPI: Nonresidential Bldg Maint & Repair', 'color': '#14b8a6', 'category': 'fire_allied'},
    'us_awe_specialty_trade': {'bls_id': 'CES2023800011', 'label': 'US AWE: Specialty Trade Contractors', 'color': '#2dd4bf', 'category': 'fire_allied'},
    # Auto Physical Damage (CPI + PPI + AWE) — 3 separate CPI vehicle series
    'us_vehicle_servicing': {'bls_id': 'CUUR0000SETD02', 'label': 'US CPI: Vehicle Maintenance & Servicing', 'color': '#0e7490', 'category': 'auto_physical'},
    'us_vehicle_repair':    {'bls_id': 'CUUR0000SETD03', 'label': 'US CPI: Vehicle Repair',                  'color': '#0891b2', 'category': 'auto_physical'},
    'us_vehicle_bodywork':  {'bls_id': 'CUUR0000SETD01', 'label': 'US CPI: Vehicle Body Work',               'color': '#06b6d4', 'category': 'auto_physical'},
    'us_ppi_auto_parts': {'bls_id': 'WPU1412',        'label': 'US PPI: Motor Vehicle Parts',            'color': '#22d3ee', 'category': 'auto_physical'},
    # Bodily Injury / Workers Comp (CES AWE by industry)
    'us_awe_total':      {'bls_id': 'CES0500000011',  'label': 'US AWE: Total Private',                  'color': '#1d4ed8', 'category': 'bodily_injury'},
    'us_awe_mfg':        {'bls_id': 'CES3000000011',  'label': 'US AWE: Manufacturing',                  'color': '#2563eb', 'category': 'bodily_injury'},
    'us_awe_construction': {'bls_id': 'CES2000000011', 'label': 'US AWE: Construction',                  'color': '#3b82f6', 'category': 'bodily_injury'},
    'us_awe_transport':  {'bls_id': 'CES4300000011',  'label': 'US AWE: Transportation & Warehousing',   'color': '#60a5fa', 'category': 'bodily_injury'},
    'us_awe_wholesale':  {'bls_id': 'CES4142000011',  'label': 'US AWE: Wholesale Trade',                'color': '#93c5fd', 'category': 'bodily_injury'},
    'us_awe_ed_health':  {'bls_id': 'CES6500000011',  'label': 'US AWE: Education & Health Services',    'color': '#a78bfa', 'category': 'bodily_injury'},
    'us_awe_leisure':    {'bls_id': 'CES7000000011',  'label': 'US AWE: Leisure & Hospitality',          'color': '#c4b5fd', 'category': 'bodily_injury'},
    # Insurance + Legal (AWE)
    'us_awe_finance':    {'bls_id': 'CES5500000011',  'label': 'US AWE: Financial Activities',           'color': '#7c3aed', 'category': 'insurance'},
    'us_awe_prof_biz':   {'bls_id': 'CES6000000011',  'label': 'US AWE: Professional & Business Svcs',   'color': '#8b5cf6', 'category': 'legal'},
    'us_awe_other_svc':  {'bls_id': 'CES8000000011',  'label': 'US AWE: Other Services',                 'color': '#ddd6fe', 'category': 'auto_physical'},
    # Legal (AWE)
    'us_awe_legal':      {'bls_id': 'CES5541100011',  'label': 'US AWE: Legal Services',                 'color': '#d97706', 'category': 'legal'},
    # Insurance (AWE)
    'us_awe_claims_adj': {'bls_id': 'CES5524200011',  'label': 'US AWE: Insurance Agencies & Brokerages','color': '#9333ea', 'category': 'insurance'},
    'us_awe_direct_ins': {'bls_id': 'CES5524120011',  'label': 'US AWE: Direct Insurance (exc Life/Health)','color': '#c084fc', 'category': 'insurance'},
    # Bodily Injury (AWE)
    'us_awe_info':       {'bls_id': 'CES5000000011',  'label': 'US AWE: Information',                    'color': '#0ea5e9', 'category': 'bodily_injury'},
    'us_awe_mining':     {'bls_id': 'CES1000000011',  'label': 'US AWE: Mining & Logging',               'color': '#475569', 'category': 'bodily_injury'},
    'us_awe_finance_bi': {'bls_id': 'CES5500000011',  'label': 'US AWE: Financial Activities',           'color': '#6366f1', 'category': 'bodily_injury'},
}

# ── Category definitions ─────────────────────────────────────────────────────
CATEGORIES = {
    'medical':        {'label': 'Medical',              'series': []},
    'legal':          {'label': 'Legal',                'series': []},
    'insurance':      {'label': 'Insurance',            'series': []},
    'bodily_injury':  {'label': 'Bodily Injury',        'series': []},
    'fire_allied':    {'label': 'Fire & Allied',        'series': []},
    'auto_physical':  {'label': 'Auto Physical Damage', 'series': []},
}

# Build category → series lists from all definitions
def _build_category_map():
    cats = {k: {'label': v['label'], 'series': []} for k, v in CATEGORIES.items()}
    for key, info in {**ONS_SERIES, **EU_HICP_SERIES, **EU_LEGAL_NL, **EU_LEGAL_IT, **EU_LEGAL_DE, **EU_LEGAL_FR, **EU_PPI_SERIES, **EU_CONSTRUCTION, **EU_LCI, **BLS_SERIES}.items():
        cat = info.get('category')
        if cat and cat in cats:
            cats[cat]['series'].append(key)
    return cats

CATEGORY_MAP = _build_category_map()

MONTH_MAP = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
    'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12,
}


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — ONS fetching
# ══════════════════════════════════════════════════════════════════════════════

def _parse_ons_json(data):
    """Parse ONS timeseries JSON. Returns list of {year, month, value, date}."""
    points = []
    cutoff = datetime.utcnow().year - 25

    for entry in data.get('months', []):
        month_name = entry.get('month', '')
        if month_name not in MONTH_MAP:
            continue
        try:
            year = int(entry.get('year', ''))
            value = float(entry.get('value', ''))
        except (ValueError, TypeError):
            continue
        if year < cutoff:
            continue
        month = MONTH_MAP[month_name]
        points.append({'year': year, 'month': month, 'value': value, 'date': f'{year}-{str(month).zfill(2)}'})

    # Also parse quarterly data for SPPI
    for entry in data.get('quarters', []):
        q = entry.get('quarter', '')
        q_map = {'Q1': 3, 'Q2': 6, 'Q3': 9, 'Q4': 12}
        if q not in q_map:
            continue
        try:
            year = int(entry.get('year', ''))
            value = float(entry.get('value', ''))
        except (ValueError, TypeError):
            continue
        if year < cutoff:
            continue
        month = q_map[q]
        points.append({'year': year, 'month': month, 'value': value, 'date': f'{year}-{q}', 'quarter': q})

    points.sort(key=lambda p: (p['year'], p['month']))
    return points


def _compute_yoy(points):
    """Convert index-level points to YoY % change."""
    by_date = {(p['year'], p['month']): p['value'] for p in points}
    result = []
    for p in points:
        prior_key = (p['year'] - 1, p['month'])
        prior_val = by_date.get(prior_key)
        if prior_val is not None and prior_val != 0:
            yoy = ((p['value'] - prior_val) / abs(prior_val)) * 100
            out = {'year': p['year'], 'month': p['month'], 'value': round(yoy, 2), 'date': p['date']}
            if 'quarter' in p:
                out['quarter'] = p['quarter']
            result.append(out)
    return result


def _fetch_ons_series():
    """Fetch all ONS insurance series. Returns (yoy_dict, raw_dict)."""
    yoy_data = {}
    raw_data = {}
    headers = {'User-Agent': USER_AGENT}

    for key, info in ONS_SERIES.items():
        url_template = ONS_URLS.get(info['dataset'])
        if not url_template:
            continue
        url = url_template.format(cdid=info['cdid'])

        try:
            resp = requests.get(url, timeout=30, headers=headers)
            if resp.status_code != 200:
                logger.warning(f"ONS {resp.status_code} for {key} ({info['cdid']})")
                continue
            points = _parse_ons_json(resp.json())
            if points:
                raw_data[key] = points
                yoy_data[key] = _compute_yoy(points)
        except Exception as e:
            logger.warning(f"ONS fetch failed for {key}: {e}")

    return yoy_data, raw_data


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2b — US BLS fetching (CPI-U, PPI, CES/AWE)
# ══════════════════════════════════════════════════════════════════════════════

BLS_PERIOD_MAP = {f'M{str(i).zfill(2)}': i for i in range(1, 13)}


def _fetch_bls_series():
    """Fetch all US BLS insurance series via API v2. Returns (yoy_dict, raw_dict).
    Makes multiple requests to get maximum historical data (BLS allows 20yr per request with key).
    """
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    from config import Config
    api_key = Config.BLS_API_KEY
    current_year = datetime.utcnow().year

    # Build list of all BLS series IDs (deduplicated for API request)
    series_ids = list(set(info['bls_id'] for info in BLS_SERIES.values()))
    id_to_keys = {}
    for key, info in BLS_SERIES.items():
        id_to_keys.setdefault(info['bls_id'], []).append(key)

    # BLS API allows max 20-year range per request with key, 10 without
    # Make multiple requests to go back as far as possible
    if api_key:
        ranges = [(2000, 2019), (2020, current_year)]
    else:
        ranges = [(current_year - 10, current_year)]

    # Accumulate points per series across all requests
    all_points = {key: [] for key in BLS_SERIES}

    for start_yr, end_yr in ranges:
        payload = {
            'seriesid': series_ids,
            'startyear': str(start_yr),
            'endyear': str(end_yr),
        }
        if api_key:
            payload['registrationkey'] = api_key

        try:
            resp = requests.post(BLS_API_URL, json=payload,
                                 headers={'Content-Type': 'application/json'},
                                 timeout=60, verify=False)
            resp.raise_for_status()
            result = resp.json()

            if result.get('status') != 'REQUEST_SUCCEEDED':
                logger.warning(f"BLS API error ({start_yr}-{end_yr}): {result.get('message', 'Unknown')}")
                continue

            for series in result.get('Results', {}).get('series', []):
                series_id = series.get('seriesID', '')
                keys = id_to_keys.get(series_id, [])
                if not keys:
                    continue

                for item in series.get('data', []):
                    period = item.get('period', '')
                    if period not in BLS_PERIOD_MAP:
                        continue
                    value = item.get('value', '')
                    if value == '-' or value == '':
                        continue
                    try:
                        point = {
                            'year': int(item.get('year', '')),
                            'month': BLS_PERIOD_MAP[period],
                            'value': float(value),
                            'date': f'{item["year"]}-{str(BLS_PERIOD_MAP[period]).zfill(2)}',
                        }
                        for key in keys:
                            all_points[key].append(dict(point))
                    except (ValueError, TypeError):
                        continue

        except Exception as e:
            logger.warning(f"BLS fetch failed ({start_yr}-{end_yr}): {e}")

    # Deduplicate and sort
    yoy_data = {}
    raw_data = {}
    for key, points in all_points.items():
        # Deduplicate by (year, month)
        seen = set()
        unique = []
        for p in points:
            k = (p['year'], p['month'])
            if k not in seen:
                seen.add(k)
                unique.append(p)
        unique.sort(key=lambda p: (p['year'], p['month']))
        if unique:
            raw_data[key] = unique
            yoy_data[key] = _compute_yoy(unique)

    logger.info(f"BLS: fetched {len(raw_data)} series with data")
    return yoy_data, raw_data


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — Eurostat fetching
# ══════════════════════════════════════════════════════════════════════════════

def _parse_eurostat_jsonstat(resp_json, series_map, dimension_key):
    """Parse Eurostat JSON-stat 2.0 response. Returns {key: [{year,month,value,date}]}."""
    dims = resp_json.get('dimension', {})
    values = resp_json.get('value', {})
    if not values:
        return {}

    time_dim = dims.get('time', {}).get('category', {}).get('index', {})
    code_dim = dims.get(dimension_key, {}).get('category', {}).get('index', {})
    if not time_dim or not code_dim:
        return {}

    num_times = len(time_dim)

    # Build reverse map: code → [series_key, ...] (supports duplicate codes)
    code_to_keys = {}
    for key, info in series_map.items():
        code = info.get('coicop') or info.get('nace') or info.get('indic')
        if code:
            code_to_keys.setdefault(code, []).append(key)

    result = {}
    for code, code_pos in code_dim.items():
        series_keys = code_to_keys.get(code, [])
        if not series_keys:
            continue

        points = []
        for period, time_pos in time_dim.items():
            idx = str(code_pos * num_times + time_pos)
            val = values.get(idx)
            if val is None:
                continue

            try:
                parts = period.split('-')
                year = int(parts[0])
                if parts[1].startswith('Q'):
                    quarter = parts[1]
                    q_num = int(quarter[1])
                    month = q_num * 3
                    date_str = f'{year}-{quarter}'
                    points.append({'year': year, 'month': month, 'quarter': quarter, 'value': float(val), 'date': date_str})
                else:
                    month = int(parts[1])
                    points.append({'year': year, 'month': month, 'value': float(val), 'date': f'{year}-{str(month).zfill(2)}'})
            except (ValueError, IndexError):
                continue

        points.sort(key=lambda p: (p['year'], p['month']))
        if points:
            for series_key in series_keys:
                result[series_key] = [dict(p) for p in points]

    return result


def _fetch_eurostat_hicp():
    """Fetch EU27 HICP series (annual rates — no YoY needed)."""
    codes = '&'.join(f'coicop={info["coicop"]}' for info in EU_HICP_SERIES.values())
    url = f'{EUROSTAT_BASE}/prc_hicp_manr?geo=EU27_2020&unit=RCH_A&freq=M&sinceTimePeriod=2000-01&{codes}'
    try:
        resp = requests.get(url, timeout=60, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), EU_HICP_SERIES, 'coicop')
    except Exception as e:
        logger.warning(f"Eurostat HICP fetch failed: {e}")
    return {}


def _fetch_eurostat_legal(geo, series_map):
    """Fetch country-specific HICP legal proxy series."""
    codes = '&'.join(f'coicop={info["coicop"]}' for info in series_map.values())
    url = f'{EUROSTAT_BASE}/prc_hicp_manr?geo={geo}&unit=RCH_A&freq=M&sinceTimePeriod=2000-01&{codes}'
    try:
        resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), series_map, 'coicop')
    except Exception as e:
        logger.warning(f"Eurostat legal ({geo}) fetch failed: {e}")
    return {}


def _fetch_eurostat_ppi():
    """Fetch EU27 PPI series (monthly, YoY rate)."""
    result = {}
    for key, info in EU_PPI_SERIES.items():
        url = f'{EUROSTAT_BASE}/{info["dataset"]}?geo=EU27_2020&nace_r2={info["nace"]}&s_adj=NSA&unit=PCH_PRE&freq=M&sinceTimePeriod=2000-01'
        try:
            resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
            if resp.status_code == 200:
                parsed = _parse_eurostat_jsonstat(resp.json(), {key: info}, 'nace_r2')
                result.update(parsed)
        except Exception as e:
            logger.warning(f"Eurostat PPI ({key}) fetch failed: {e}")
    return result


def _fetch_eurostat_construction():
    """Fetch EU27 construction price indices (quarterly)."""
    indics = '&'.join(f'indic_bt={info["indic"]}' for info in EU_CONSTRUCTION.values())
    url = f'{EUROSTAT_BASE}/sts_copi_q?geo=EU27_2020&s_adj=NSA&unit=PCH_Q4&freq=Q&sinceTimePeriod=2000-Q1&{indics}'
    try:
        resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), EU_CONSTRUCTION, 'indic_bt')
    except Exception as e:
        logger.warning(f"Eurostat construction fetch failed: {e}")
    return {}


def _fetch_eurostat_lci():
    """Fetch EU27 Labour Cost Index (quarterly)."""
    unique_naces = sorted(set(info['nace'] for info in EU_LCI.values()))
    naces = '&'.join(f'nace_r2={n}' for n in unique_naces)
    url = f'{EUROSTAT_BASE}/lc_lci_r2_q?geo=EU27_2020&s_adj=SCA&lcstruct=D1_D4_MD5&unit=PCH_Q4&freq=Q&sinceTimePeriod=2000-Q1&{naces}'
    try:
        resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), EU_LCI, 'nace_r2')
    except Exception as e:
        logger.warning(f"Eurostat LCI fetch failed: {e}")
    return {}


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — ONS Construction OPI (Excel download)
# ══════════════════════════════════════════════════════════════════════════════

CONSTRUCTION_OPI_URL = 'https://www.ons.gov.uk/file?uri=/businessindustryandtrade/constructionindustry/datasets/interimconstructionoutputpriceindices/current/bulletindataset9.xlsx'

def _fetch_construction_opi():
    """Download ONS Construction OPI Excel and extract new work + R&M indices."""
    try:
        from openpyxl import load_workbook

        resp = requests.get(CONSTRUCTION_OPI_URL, timeout=60, headers={'User-Agent': USER_AGENT})
        if resp.status_code != 200:
            logger.warning(f"Construction OPI download failed: {resp.status_code}")
            return {}, {}

        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
        result = {}

        # Try to find the data sheet — ONS names vary
        target_sheet = None
        for name in wb.sheetnames:
            if 'data' in name.lower() or 'table' in name.lower() or 'index' in name.lower():
                target_sheet = wb[name]
                break
        if not target_sheet:
            target_sheet = wb[wb.sheetnames[0]]

        # Parse: look for "All New Work" and "All Repair and Maintenance" columns
        # This is fragile — ONS may change layout. We do best-effort.
        ws = target_sheet
        header_row = None
        new_col = None
        repair_col = None

        for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
            for cell in row:
                val = str(cell.value or '').lower()
                if 'all new work' in val:
                    new_col = cell.column
                    header_row = cell.row
                elif 'repair' in val and 'maintenance' in val:
                    repair_col = cell.column
                    header_row = cell.row

        if not header_row:
            logger.warning("Construction OPI: could not find header row")
            return {}, {}

        # Parse data rows (quarterly: "2024 Q1" format in first column)
        new_points = []
        repair_points = []
        q_map = {'Q1': 3, 'Q2': 6, 'Q3': 9, 'Q4': 12}

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
            period_val = str(row[0].value or '')
            parts = period_val.strip().split()
            if len(parts) < 2 or parts[1] not in q_map:
                continue
            try:
                year = int(parts[0])
                quarter = parts[1]
                month = q_map[quarter]
            except ValueError:
                continue

            if new_col:
                try:
                    val = float(row[new_col - 1].value)
                    new_points.append({'year': year, 'month': month, 'quarter': quarter, 'value': val, 'date': f'{year}-{quarter}'})
                except (ValueError, TypeError):
                    pass
            if repair_col:
                try:
                    val = float(row[repair_col - 1].value)
                    repair_points.append({'year': year, 'month': month, 'quarter': quarter, 'value': val, 'date': f'{year}-{quarter}'})
                except (ValueError, TypeError):
                    pass

        yoy_result = {}
        raw_result = {}
        if new_points:
            raw_result['uk_opi_new'] = new_points
            yoy_result['uk_opi_new'] = _compute_yoy(new_points)
        if repair_points:
            raw_result['uk_opi_repair'] = repair_points
            yoy_result['uk_opi_repair'] = _compute_yoy(repair_points)

        wb.close()
        return yoy_result, raw_result

    except Exception as e:
        logger.warning(f"Construction OPI fetch failed: {e}")
        return {}, {}


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4b — Eurostat index fetchers (raw indices for QoQ computation)
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_eurostat_hicp_index():
    """Fetch EU27 HICP monthly index (2015=100)."""
    codes = '&'.join(f'coicop={info["coicop"]}' for info in EU_HICP_SERIES.values())
    url = f'{EUROSTAT_BASE}/prc_hicp_midx?geo=EU27_2020&unit=I15&freq=M&sinceTimePeriod=2000-01&{codes}'
    try:
        resp = requests.get(url, timeout=60, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), EU_HICP_SERIES, 'coicop')
    except Exception as e:
        logger.warning(f"Eurostat HICP index fetch failed: {e}")
    return {}


def _fetch_eurostat_legal_index(geo, series_map):
    """Fetch country-specific HICP legal proxy index."""
    codes = '&'.join(f'coicop={info["coicop"]}' for info in series_map.values())
    url = f'{EUROSTAT_BASE}/prc_hicp_midx?geo={geo}&unit=I15&freq=M&sinceTimePeriod=2000-01&{codes}'
    try:
        resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), series_map, 'coicop')
    except Exception as e:
        logger.warning(f"Eurostat legal index ({geo}) fetch failed: {e}")
    return {}


def _fetch_eurostat_ppi_index():
    """Fetch EU27 PPI monthly index (2015=100)."""
    result = {}
    for key, info in EU_PPI_SERIES.items():
        url = f'{EUROSTAT_BASE}/{info["dataset"]}?geo=EU27_2020&nace_r2={info["nace"]}&s_adj=NSA&unit=I15&freq=M&sinceTimePeriod=2000-01'
        try:
            resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
            if resp.status_code == 200:
                parsed = _parse_eurostat_jsonstat(resp.json(), {key: info}, 'nace_r2')
                result.update(parsed)
        except Exception as e:
            logger.warning(f"Eurostat PPI index ({key}) fetch failed: {e}")
    return result


def _fetch_eurostat_construction_index():
    """Fetch EU27 construction quarterly index (2020=100)."""
    indics = '&'.join(f'indic_bt={info["indic"]}' for info in EU_CONSTRUCTION.values())
    url = f'{EUROSTAT_BASE}/sts_copi_q?geo=EU27_2020&s_adj=NSA&unit=I20&freq=Q&sinceTimePeriod=2000-Q1&{indics}'
    try:
        resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), EU_CONSTRUCTION, 'indic_bt')
    except Exception as e:
        logger.warning(f"Eurostat construction index fetch failed: {e}")
    return {}


def _fetch_eurostat_lci_index():
    """Fetch EU27 LCI quarterly index (2020=100)."""
    unique_naces = sorted(set(info['nace'] for info in EU_LCI.values()))
    naces = '&'.join(f'nace_r2={n}' for n in unique_naces)
    url = f'{EUROSTAT_BASE}/lc_lci_r2_q?geo=EU27_2020&s_adj=SCA&lcstruct=D1_D4_MD5&unit=I20&freq=Q&sinceTimePeriod=2000-Q1&{naces}'
    try:
        resp = requests.get(url, timeout=30, headers={'User-Agent': USER_AGENT})
        if resp.status_code == 200:
            return _parse_eurostat_jsonstat(resp.json(), EU_LCI, 'nace_r2')
    except Exception as e:
        logger.warning(f"Eurostat LCI index fetch failed: {e}")
    return {}


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4c — Shared helpers (QoQ computation, quarterly aggregation)
# ══════════════════════════════════════════════════════════════════════════════

def aggregate_monthly_to_quarterly(points):
    """Average monthly values into quarterly buckets. Works for both rates and indices."""
    from collections import defaultdict
    buckets = defaultdict(list)
    for p in points:
        q = (p['month'] - 1) // 3 + 1
        key = (p['year'], q)
        buckets[key].append(p['value'])
    result = []
    for (year, q), vals in sorted(buckets.items()):
        avg = sum(vals) / len(vals)
        result.append({
            'year': year, 'month': q * 3, 'quarter': f'Q{q}',
            'value': round(avg, 4), 'date': f'{year}-Q{q}'
        })
    return result


def compute_qoq(raw_points, is_quarterly_native=False):
    """Compute QoQ % change from raw index points.
    Monthly → aggregate to quarterly first, then QoQ.
    Quarterly → QoQ directly.
    """
    if not raw_points:
        return []

    if is_quarterly_native:
        points = sorted(raw_points, key=lambda p: (p['year'], p['month']))
    else:
        points = aggregate_monthly_to_quarterly(raw_points)

    # Build lookup by (year, quarter_num)
    by_q = {}
    for p in points:
        q = int(p.get('quarter', f'Q{(p["month"] - 1) // 3 + 1}').replace('Q', ''))
        by_q[(p['year'], q)] = p['value']

    result = []
    for p in points:
        q = int(p.get('quarter', f'Q{(p["month"] - 1) // 3 + 1}').replace('Q', ''))
        prev_q = q - 1
        prev_year = p['year']
        if prev_q == 0:
            prev_q = 4
            prev_year -= 1
        prev_val = by_q.get((prev_year, prev_q))
        if prev_val is not None and prev_val != 0:
            qoq = ((p['value'] - prev_val) / abs(prev_val)) * 100
            result.append({
                'year': p['year'], 'month': p['month'],
                'quarter': p.get('quarter', f'Q{q}'),
                'value': round(qoq, 2),
                'date': p.get('date', f'{p["year"]}-Q{q}'),
            })
    return result


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — Assemble all data
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_all():
    """Fetch all insurance inflation series from ONS + Eurostat + BLS."""
    all_series = {}
    all_series_raw = {}
    series_meta = {}

    # ONS UK series (returns yoy + raw)
    logger.info("Fetching ONS insurance series...")
    ons_yoy, ons_raw = _fetch_ons_series()
    all_series.update(ons_yoy)
    all_series_raw.update(ons_raw)

    # ONS Construction OPI (returns yoy + raw)
    logger.info("Fetching ONS Construction OPI...")
    opi_yoy, opi_raw = _fetch_construction_opi()
    all_series.update(opi_yoy)
    all_series_raw.update(opi_raw)

    # US BLS series (CPI-U, PPI, CES/AWE)
    logger.info("Fetching US BLS insurance series...")
    bls_yoy, bls_raw = _fetch_bls_series()
    all_series.update(bls_yoy)
    all_series_raw.update(bls_raw)

    # Eurostat HICP (rates)
    logger.info("Fetching Eurostat HICP...")
    all_series.update(_fetch_eurostat_hicp())

    # Eurostat HICP (indices for QoQ)
    logger.info("Fetching Eurostat HICP indices...")
    all_series_raw.update(_fetch_eurostat_hicp_index())

    # Eurostat NL/IT/DE/FR legal (rates + indices)
    logger.info("Fetching Eurostat legal proxies (NL/IT/DE/FR)...")
    all_series.update(_fetch_eurostat_legal('NL', EU_LEGAL_NL))
    all_series.update(_fetch_eurostat_legal('IT', EU_LEGAL_IT))
    all_series.update(_fetch_eurostat_legal('DE', EU_LEGAL_DE))
    all_series.update(_fetch_eurostat_legal('FR', EU_LEGAL_FR))
    all_series_raw.update(_fetch_eurostat_legal_index('NL', EU_LEGAL_NL))
    all_series_raw.update(_fetch_eurostat_legal_index('IT', EU_LEGAL_IT))
    all_series_raw.update(_fetch_eurostat_legal_index('DE', EU_LEGAL_DE))
    all_series_raw.update(_fetch_eurostat_legal_index('FR', EU_LEGAL_FR))

    # Eurostat PPI (rates + indices)
    logger.info("Fetching Eurostat PPI...")
    all_series.update(_fetch_eurostat_ppi())
    all_series_raw.update(_fetch_eurostat_ppi_index())

    # Eurostat Construction (rates + indices)
    logger.info("Fetching Eurostat Construction...")
    all_series.update(_fetch_eurostat_construction())
    all_series_raw.update(_fetch_eurostat_construction_index())

    # Eurostat LCI (rates + indices)
    logger.info("Fetching Eurostat LCI...")
    all_series.update(_fetch_eurostat_lci())
    all_series_raw.update(_fetch_eurostat_lci_index())

    # Build series metadata
    all_defs = {**ONS_SERIES, **EU_HICP_SERIES, **EU_LEGAL_NL, **EU_LEGAL_IT, **EU_LEGAL_DE, **EU_LEGAL_FR, **EU_PPI_SERIES, **BLS_SERIES}
    # Add Construction OPI defs
    all_defs['uk_opi_new'] = {'label': 'UK Construction OPI: New Work', 'color': '#f97316', 'category': 'fire_allied'}
    all_defs['uk_opi_repair'] = {'label': 'UK Construction OPI: Repair & Maintenance', 'color': '#fb923c', 'category': 'fire_allied'}
    # Add EU construction and LCI
    for key, info in EU_CONSTRUCTION.items():
        all_defs[key] = info
    for key, info in EU_LCI.items():
        all_defs[key] = info

    for key, info in all_defs.items():
        source = 'BLS' if key.startswith('us_') else ('ONS' if key.startswith('uk_') else 'Eurostat')
        freq = 'Q' if key in EU_CONSTRUCTION or key in EU_LCI or key in ('uk_opi_new', 'uk_opi_repair') or key == 'uk_sppi_legal' else 'M'
        meta = {
            'label': info.get('label', key),
            'color': info.get('color', '#94a3b8'),
            'source': source,
            'freq': freq,
        }
        if info.get('approximate'):
            meta['approximate'] = True
        series_meta[key] = meta

    # Update category map with construction OPI
    cats = {k: dict(v) for k, v in CATEGORY_MAP.items()}
    if 'uk_opi_new' in all_series:
        cats['fire_allied']['series'].append('uk_opi_new')
    if 'uk_opi_repair' in all_series:
        cats['fire_allied']['series'].append('uk_opi_repair')

    # Filter to only include series that actually have data
    for cat_key, cat_info in cats.items():
        cat_info['series'] = [s for s in cat_info['series'] if s in all_series]

    return {
        'series': all_series,
        'series_raw': all_series_raw,
        'categories': cats,
        'series_meta': series_meta,
        'meta': {
            'source': 'BLS, ONS, Eurostat',
            'description': 'Insurance/Reinsurance Inflation Indicators',
            'frequency': 'Monthly & Quarterly',
            'total_series': len(all_series),
            'total_series_raw': len(all_series_raw),
            'last_updated': datetime.utcnow().isoformat(),
        }
    }


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — Cache and public API
# ══════════════════════════════════════════════════════════════════════════════

class InsuranceInflationCache:
    def __init__(self):
        self._lock = threading.RLock()
        self._data = None
        self._last_fetch = 0
        self._last_fail = 0

    def get(self):
        with self._lock:
            if self._data and (time.time() - self._last_fetch) < CACHE_TTL:
                return self._data
            if self._last_fail and (time.time() - self._last_fail) < RETRY_BACKOFF:
                return self._data or _empty_result()
        data = _fetch_all()
        if data and data.get('series'):
            with self._lock:
                self._data = data
                self._last_fetch = time.time()
                self._last_fail = 0
            return data
        with self._lock:
            self._last_fail = time.time()
            return self._data or _empty_result()


_cache = InsuranceInflationCache()


def _empty_result():
    return {
        'series': {},
        'series_raw': {},
        'categories': CATEGORY_MAP,
        'series_meta': {},
        'meta': {'source': 'BLS, ONS, Eurostat', 'error': 'No data available'},
    }


def get_insurance_inflation_data():
    """Public API: returns cached insurance inflation data."""
    return _cache.get()
