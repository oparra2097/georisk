from io import BytesIO
from datetime import datetime
from flask import Blueprint, jsonify, send_file, request
from backend.cache.store import store
from backend.data_sources.market_data import get_market_data, get_market_history
from backend.data_sources.imf_cofer import (
    get_cofer_data,
    refresh_cache as refresh_cofer_cache,
    diagnose_fetch as diagnose_cofer_fetch,
)
from backend.data_sources.reserves_nowcast import get_nowcast_data
from backend.data_sources.bls_cpi import get_bls_cpi_data, get_bls_components, clear_bls_caches
from backend.data_sources.ons_cpi import get_ons_cpi_data, get_ons_components
from backend.data_sources.eurostat_hicp import get_eurostat_cpi_data, get_eurostat_components
from backend.data_sources.substack_feed import get_substack_posts
from backend.data_sources.commodities_forecast import get_forecast_data
from backend.data_sources.gdp_nowcast import get_gdp_nowcast
from backend.data_sources.imf_weo import get_weo_data
from backend.data_sources.world_bank import get_wb_data
from backend.data_sources.sovereign_debt import get_sovereign_debt_data
from backend.data_sources.fertilizer_em_inflation import get_fertilizer_em_data
from backend.data_sources.yale_tariff import get_yale_tariff_data
from backend.data_sources.insurance_inflation import get_insurance_inflation_data
from backend.data_sources.em_vulnerability import get_em_vulnerability_data
from flask_login import login_required, current_user
from functools import wraps
from backend.cache.database import get_country_history, get_all_history, detect_anomalies, get_score_count
from config import Config

api_bp = Blueprint('api', __name__)


def insurance_access_required(f):
    """Decorator: requires verified @aig.com email or admin-granted access."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'error': 'Authentication required', 'login_url': '/auth/login'}), 401
        if not current_user.email_verified:
            return jsonify({'error': 'Please verify your email address first. Check your inbox for the verification link.'}), 403
        if not current_user.has_insurance_access():
            return jsonify({'error': 'Insurance data access requires an @aig.com email address. Contact the administrator for access.'}), 403
        return f(*args, **kwargs)
    return decorated


@api_bp.route('/scores')
def get_all_scores():
    """Return risk scores for all countries."""
    all_scores = store.get_all_scores()
    result = {}
    for code, risk in all_scores.items():
        result[code] = risk.to_dict()
    return jsonify(result)


@api_bp.route('/scores/<country_code>')
def get_country_score(country_code):
    """Return detailed score for a single country."""
    code = country_code.upper()
    risk = store.get_country(code)
    if not risk:
        return jsonify({'error': 'Country not found'}), 404
    return jsonify(risk.to_dict())


@api_bp.route('/headlines/<country_code>')
def get_country_headlines(country_code):
    """Return recent headlines for a country."""
    code = country_code.upper()
    if code == 'GLOBAL':
        articles = store.get_global_headlines()
    else:
        articles = store.get_headlines(code)
    return jsonify({
        'articles': [a.to_dict() for a in articles]
    })


@api_bp.route('/headlines/global')
def get_global_headlines():
    """Return global geopolitical headlines."""
    articles = store.get_global_headlines()
    return jsonify({
        'articles': [a.to_dict() for a in articles]
    })


@api_bp.route('/hotspots')
def get_hotspots():
    """Return countries with risk score above threshold."""
    threshold = Config.HOTSPOT_THRESHOLD
    hotspots = store.get_hotspots(threshold)
    hotspots.sort(key=lambda x: x.composite_score, reverse=True)
    return jsonify({
        'hotspots': [h.to_dict() for h in hotspots]
    })


@api_bp.route('/export')
def export_excel():
    """Generate and return an Excel file with all country risk scores."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    all_scores = store.get_all_scores()
    rows = sorted(all_scores.values(), key=lambda r: r.composite_score, reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'GeoRisk Scores'

    # Header row
    headers = [
        'Country', 'Code', 'Composite Score',
        'Political Stability', 'Military Conflict', 'Economic Sanctions',
        'Protests/Civil Unrest', 'Terrorism', 'Diplomatic Tensions',
        'Avg Tone', 'GDELT Events', 'Updated At'
    ]
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    thin_border = Border(
        bottom=Side(style='thin', color='374151')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, risk in enumerate(rows, 2):
        ind = risk.indicators
        ws.cell(row=row_idx, column=1, value=risk.country_name)
        ws.cell(row=row_idx, column=2, value=risk.country_code)
        ws.cell(row=row_idx, column=3, value=round(risk.composite_score, 1))
        ws.cell(row=row_idx, column=4, value=round(ind.political_stability, 1))
        ws.cell(row=row_idx, column=5, value=round(ind.military_conflict, 1))
        ws.cell(row=row_idx, column=6, value=round(ind.economic_sanctions, 1))
        ws.cell(row=row_idx, column=7, value=round(ind.protests_civil_unrest, 1))
        ws.cell(row=row_idx, column=8, value=round(ind.terrorism, 1))
        ws.cell(row=row_idx, column=9, value=round(ind.diplomatic_tensions, 1))
        ws.cell(row=row_idx, column=10, value=round(risk.avg_tone, 2))
        ws.cell(row=row_idx, column=11, value=risk.gdelt_event_count)
        ws.cell(row=row_idx, column=12, value=risk.updated_at.strftime('%Y-%m-%d %H:%M') if risk.updated_at else '')

        # Color composite score cell based on risk level
        score = risk.composite_score
        if score >= 70:
            ws.cell(row=row_idx, column=3).fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        elif score >= 40:
            ws.cell(row=row_idx, column=3).fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for row in range(2, min(len(rows) + 2, 50)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Write to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'georisk_scores_{today}.xlsx'
    )


@api_bp.route('/markets')
def get_markets():
    """Return live market data (cached 5 minutes)."""
    data = get_market_data()
    return jsonify(data)


@api_bp.route('/markets/history')
def get_markets_history():
    """Return historical price data for a symbol and time period."""
    symbol = request.args.get('symbol', '')
    period = request.args.get('period', '1mo')

    if not symbol:
        return jsonify({'error': 'symbol parameter required'}), 400

    valid_periods = ['1d', '5d', '1mo', '1y', '5y', '10y']
    if period not in valid_periods:
        return jsonify({'error': f'Invalid period. Use: {", ".join(valid_periods)}'}), 400

    data = get_market_history(symbol, period)
    if data is None:
        return jsonify({'error': 'No data available'}), 404

    return jsonify(data)


@api_bp.route('/cofer')
def get_cofer():
    """Return central bank reserves data (cached 24 hours)."""
    data = get_cofer_data()
    return jsonify(data)


@api_bp.route('/cofer/refresh', methods=['POST', 'GET'])
def refresh_cofer():
    """Force-clear the reserves cache and return a detailed diagnostic.

    Runs the full fetch chain (IMF Data API → DBnomics → World Bank)
    with per-attempt logging, stores the result in the cache, and
    returns a JSON diagnostic so you can see exactly which endpoint
    served the data (and why the others failed). Useful when the
    upstream data provider releases a new period and you don't want
    to wait for the 24-hour TTL to expire.
    """
    diagnostic = diagnose_cofer_fetch()
    return jsonify({'ok': bool(diagnostic.get('source')), **diagnostic})


@api_bp.route('/cofer/nowcast')
def get_cofer_nowcast():
    """Return reserves nowcast — real-time currency composition estimates."""
    data = get_nowcast_data()
    return jsonify(data)


@api_bp.route('/cpi/us')
def get_us_cpi():
    """Return US CPI data from BLS (cached 24 hours)."""
    data = get_bls_cpi_data()
    return jsonify(data)


@api_bp.route('/cpi/uk')
def get_uk_cpi():
    """Return UK CPI data from ONS (cached 24 hours)."""
    data = get_ons_cpi_data()
    return jsonify(data)


@api_bp.route('/cpi/us/components')
def get_us_cpi_components():
    """Return US CPI component breakdown from BLS (cached 24 hours)."""
    data = get_bls_components()
    return jsonify(data)


@api_bp.route('/cpi/uk/components')
def get_uk_cpi_components():
    """Return UK CPI component breakdown from ONS (cached 24 hours)."""
    data = get_ons_components()
    return jsonify(data)


@api_bp.route('/cpi/us/refresh', methods=['POST'])
def refresh_us_cpi():
    """Clear BLS caches and force re-fetch on next request."""
    clear_bls_caches()
    return jsonify({'status': 'ok', 'message': 'BLS CPI caches cleared'})


@api_bp.route('/cpi/us/export')
def export_us_cpi_excel():
    """Generate Excel file with US CPI overview data."""
    return _export_cpi_excel(get_bls_cpi_data(), 'US CPI', 'us_cpi_data')


@api_bp.route('/cpi/uk/export')
def export_uk_cpi_excel():
    """Generate Excel file with UK CPI overview data."""
    return _export_cpi_excel(get_ons_cpi_data(), 'UK CPI', 'uk_cpi_data')


@api_bp.route('/cpi/us/components/export')
def export_us_components_excel():
    """Generate Excel file with US CPI component breakdown."""
    return _export_cpi_excel(get_bls_components(), 'US CPI Components', 'us_cpi_components')


@api_bp.route('/cpi/uk/components/export')
def export_uk_components_excel():
    """Generate Excel file with UK CPI component breakdown."""
    return _export_cpi_excel(get_ons_components(), 'UK CPI Components', 'uk_cpi_components')


@api_bp.route('/cpi/eu')
def get_eu_cpi():
    """Return EU HICP data from Eurostat (cached 24 hours)."""
    data = get_eurostat_cpi_data()
    return jsonify(data)


@api_bp.route('/cpi/eu/components')
def get_eu_cpi_components():
    """Return EU HICP component breakdown from Eurostat (cached 24 hours)."""
    data = get_eurostat_components()
    return jsonify(data)


@api_bp.route('/cpi/eu/export')
def export_eu_cpi_excel():
    """Generate Excel file with EU HICP overview data."""
    return _export_cpi_excel(get_eurostat_cpi_data(), 'EU HICP', 'eu_hicp_data')


@api_bp.route('/cpi/eu/components/export')
def export_eu_components_excel():
    """Generate Excel file with EU HICP component breakdown."""
    return _export_cpi_excel(get_eurostat_components(), 'EU HICP Components', 'eu_hicp_components')


def _export_cpi_excel(data, title, filename_prefix):
    """Shared helper: export CPI data (overview or components) to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    series = data.get('series', {})
    categories = data.get('categories', {})
    meta = data.get('meta', {})

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    wb = Workbook()
    first_sheet = True

    for key, label in categories.items():
        points = series.get(key, [])
        if not points:
            continue

        if first_sheet:
            ws = wb.active
            ws.title = label[:31]
            first_sheet = False
        else:
            ws = wb.create_sheet(label[:31])

        # Title row
        ws.cell(row=1, column=1, value=f'{title}: {label}')
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        if meta.get('source'):
            ws.cell(row=2, column=1, value=meta['source'])
            ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='6B7280')

        # Header row
        row = 4
        headers = ['Date', 'Index Value', 'YoY Change (%)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows (most recent first)
        for pt in reversed(points):
            row += 1
            ws.cell(row=row, column=1, value=pt.get('date', '')).border = thin_border
            cell_val = ws.cell(row=row, column=2, value=pt.get('value'))
            cell_val.number_format = '#,##0.000'
            cell_val.alignment = Alignment(horizontal='center')
            cell_val.border = thin_border
            yoy = pt.get('yoy_change')
            cell_yoy = ws.cell(row=row, column=3, value=yoy)
            cell_yoy.number_format = '0.00'
            cell_yoy.alignment = Alignment(horizontal='center')
            cell_yoy.border = thin_border

        # Auto-width
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 16
        ws.freeze_panes = 'A5'

    if first_sheet:
        # No data at all — create empty sheet with message
        ws = wb.active
        ws.title = 'No Data'
        ws.cell(row=1, column=1, value='No data available. Check API key configuration.')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename_prefix}_{today}.xlsx'
    )


@api_bp.route('/markets/export')
def export_markets_excel():
    """Generate Excel file with current market data snapshot."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_market_data()
    markets = data.get('markets', [])

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    green_font = Font(color='10B981')
    red_font = Font(color='EF4444')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Market Data'

    # Title
    ws.cell(row=1, column=1, value='Parra Macro — Market Data Snapshot')
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}')
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='6B7280')

    # Headers
    headers = ['Name', 'Symbol', 'Type', 'Price', 'Change', 'Change %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, m in enumerate(markets):
        row = 5 + i
        ws.cell(row=row, column=1, value=m.get('name', '')).border = thin_border
        ws.cell(row=row, column=2, value=m.get('symbol', '')).border = thin_border
        ws.cell(row=row, column=3, value=m.get('type', '')).border = thin_border

        price_cell = ws.cell(row=row, column=4, value=m.get('price'))
        price_cell.number_format = '#,##0.00'
        price_cell.alignment = Alignment(horizontal='center')
        price_cell.border = thin_border

        change = m.get('change')
        change_cell = ws.cell(row=row, column=5, value=change)
        change_cell.number_format = '+#,##0.00;-#,##0.00;0.00'
        change_cell.alignment = Alignment(horizontal='center')
        change_cell.border = thin_border
        if change is not None:
            change_cell.font = green_font if change >= 0 else red_font

        pct = m.get('change_pct')
        pct_cell = ws.cell(row=row, column=6, value=pct)
        pct_cell.number_format = '+0.00%;-0.00%;0.00%'
        pct_cell.alignment = Alignment(horizontal='center')
        pct_cell.border = thin_border
        if pct is not None:
            pct_cell.font = green_font if pct >= 0 else red_font
            # Store as decimal for Excel percentage format
            pct_cell.value = pct / 100.0

    # Auto-width
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'market_data_{today}.xlsx'
    )


@api_bp.route('/forecasts')
def get_forecasts():
    """Return commodities forecast data (cached 24 hours)."""
    data = get_forecast_data()
    return jsonify(data)


# Map commodity display names → methodology markdown file names.
_METHODOLOGY_FILES = {
    'WTI Crude':        'wti_crude.md',
    'Brent Crude':      'brent_crude.md',
    'Natural Gas (HH)': 'natural_gas_hh.md',
    'TTF Gas':          'ttf_gas.md',
    'Gold':             'gold.md',
    'Silver':           'silver.md',
    'Platinum':         'platinum.md',
    'Copper':           'copper.md',
    'Aluminum':         'aluminum.md',
    'Cocoa':            'cocoa.md',
    'Wheat':            'wheat.md',
    'Soybeans':         'soybeans.md',
    'Coffee':           'coffee.md',
    'overview':         'README.md',
}


@api_bp.route('/forecasts/methodology/<path:commodity>')
def get_methodology(commodity):
    """Serve the per-commodity methodology markdown as plain text JSON.

    Frontend renders client-side via marked.js in a modal. `commodity` is
    the display name used in commodities_forecast.COMMODITIES (or the
    literal string 'overview' for the index / README).
    """
    import os
    fname = _METHODOLOGY_FILES.get(commodity)
    if not fname:
        return jsonify({'error': f'Unknown commodity: {commodity}'}), 404
    docs_dir = os.path.join(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__))), 'docs', 'commodities')
    path = os.path.join(docs_dir, fname)
    if not os.path.exists(path):
        return jsonify({'error': f'Methodology file missing: {fname}'}), 404
    try:
        with open(path) as f:
            markdown = f.read()
    except Exception as e:
        return jsonify({'error': f'Failed to read {fname}: {e}'}), 500
    return jsonify({
        'commodity': commodity,
        'markdown': markdown,
        'file': fname,
    })


@api_bp.route('/forecasts/shocks', defaults={'commodity': None})
@api_bp.route('/forecasts/shocks/<path:commodity>')
def get_shocks(commodity):
    """Return the scenario-shock catalogue.

    Without a commodity → ``{all: {commodity_name: [shocks...]}}`` for the
    full catalogue. With one → ``{commodity, shocks}``. Used by the
    frontend scenario builder to render sliders.
    """
    from backend.data_sources import commodity_models
    return jsonify(commodity_models.get_shocks_catalogue(commodity))


@api_bp.route('/forecasts/scenario', methods=['POST'])
def post_scenario_forecast():
    """Run a scenario-modified forecast for a single commodity.

    Body::

        {
            "commodity": "WTI Crude",
            "shocks": [{"id": "opec_production", "magnitude": -2.0}],
            "driver_shifts": {"yf:^GSPC": -0.01}
        }

    Both ``shocks`` and ``driver_shifts`` are optional — pass an empty
    dict to retrieve the unmodified base forecast (same as
    ``/api/forecasts`` per-commodity output).
    """
    from backend.data_sources import commodity_models
    payload = request.get_json(silent=True) or {}
    commodity = payload.get('commodity')
    if not commodity or commodity not in commodity_models.TICKERS:
        return jsonify({'error': f'Unknown or missing commodity: {commodity!r}'}), 400

    shocks = payload.get('shocks') or []
    driver_shifts = payload.get('driver_shifts') or {}
    if not isinstance(shocks, list) or not isinstance(driver_shifts, dict):
        return jsonify({'error': 'shocks must be a list, driver_shifts a dict'}), 400

    try:
        result = commodity_models.get_model_forecast(
            commodity,
            driver_shifts=driver_shifts,
            shocks=shocks,
        )
    except Exception as exc:
        return jsonify({'error': f'Forecast failed: {exc}'}), 500
    if result is None:
        return jsonify({'error': f'Model unavailable for {commodity}'}), 503

    return jsonify({
        'commodity': commodity,
        'forecast': result.get('forecast'),
        'summary': result.get('summary'),
        'exog_columns': result.get('exog_columns', []),
        'shocks_applied': shocks,
        'driver_shifts_applied': driver_shifts,
        'shocks_catalogue': commodity_models.SHOCKS.get(commodity, []),
    })


@api_bp.route('/gdp-nowcast')
def gdp_nowcast():
    """Return US GDP nowcast estimate (cached 6 hours)."""
    data = get_gdp_nowcast()
    return jsonify(data)


@api_bp.route('/forecasts/export')
def export_forecasts_excel():
    """Generate Excel file with commodity forecast + historical data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_forecast_data()
    groups = data.get('groups', {})
    time_ctx = data.get('time_context', {})
    labels = time_ctx.get('labels', [])
    label_types = time_ctx.get('label_types', [])
    year_end_labels = time_ctx.get('year_end_labels', [time_ctx.get('year_end_label', 'FY Avg')])
    fy_keys = ['FY'] + [f'FY{i+2}' for i in range(len(year_end_labels) - 1)]
    # Per-group scenario config — extracted per group below

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    section_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    forecast_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    current_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    num_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    wb = Workbook()
    first_sheet = True

    for group_name in ['Oil & Gas', 'Agriculture', 'Metals']:
        group = groups.get(group_name)
        if not group:
            continue

        # Per-group scenario config
        scenario_order = group.get('scenario_order', ['Actual'])
        scenario_weights = group.get('scenario_weights', {})
        scenario_labels = group.get('scenario_labels', {})
        commodities = group.get('commodities', {})

        for comm_name, comm_data in commodities.items():
            # Create a sheet per commodity
            if first_sheet:
                ws = wb.active
                ws.title = comm_name[:31]
                first_sheet = False
            else:
                ws = wb.create_sheet(comm_name[:31])

            unit = comm_data.get('unit', '')
            scenarios = comm_data.get('scenarios', {})
            historical = comm_data.get('historical', [])
            latest = comm_data.get('latest_close')

            row = 1

            # ── Title ──
            ws.cell(row=row, column=1, value=f'{comm_name} — {unit}')
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 1
            ws.cell(row=row, column=1, value=f'Group: {group_name}')
            row += 1
            if latest:
                ws.cell(row=row, column=1, value=f'Latest Close: {latest} {unit}')
            row += 1
            ws.cell(row=row, column=1, value=f'As of: {time_ctx.get("today", "")}')
            row += 2

            # ── Scenario Forecast Table ──
            ws.cell(row=row, column=1, value='SCENARIO FORECASTS')
            ws.cell(row=row, column=1).font = section_font
            ws.cell(row=row, column=1).fill = section_fill
            total_cols = len(labels) + len(year_end_labels) + 1
            for ci in range(2, total_cols + 1):
                ws.cell(row=row, column=ci).fill = section_fill
            row += 1

            # Header row
            fc_headers = ['Scenario'] + labels + year_end_labels
            for ci, h in enumerate(fc_headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row += 1

            # Data rows
            for sc in scenario_order:
                sc_data = scenarios.get(sc)
                if not sc_data:
                    continue
                ws.cell(row=row, column=1, value=sc).border = thin_border
                weight = scenario_weights.get(sc)
                if weight:
                    ws.cell(row=row, column=1, value=f'{sc} ({weight*100:.0f}%)')
                ws.cell(row=row, column=1).border = thin_border

                for ci, lbl in enumerate(labels, 2):
                    val = sc_data.get(lbl)
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    lt = label_types[ci - 2] if (ci - 2) < len(label_types) else ''
                    if lt == 'forecast':
                        cell.fill = forecast_fill
                    elif lt == 'current_q':
                        cell.fill = current_fill

                # FY columns (FY, FY2, ...)
                for fi, fy_key in enumerate(fy_keys):
                    fy_val = sc_data.get(fy_key)
                    cell = ws.cell(row=row, column=len(labels) + 2 + fi, value=fy_val)
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                row += 1

            row += 1

            # ── Scenario Descriptions ──
            for sc in scenario_order:
                if sc == 'Actual':
                    continue
                desc = scenario_labels.get(sc, '')
                w = scenario_weights.get(sc)
                w_str = f' ({w*100:.0f}%)' if w else ''
                ws.cell(row=row, column=1, value=f'{sc}{w_str}: {desc}')
                ws.cell(row=row, column=1).font = Font(italic=True, size=9, color='6B7280')
                row += 1

            row += 2

            # ── Historical Quarterly Data ──
            if historical:
                ws.cell(row=row, column=1, value='HISTORICAL QUARTERLY AVERAGES')
                ws.cell(row=row, column=1).font = section_font
                ws.cell(row=row, column=1).fill = section_fill
                for ci in range(2, 4):
                    ws.cell(row=row, column=ci).fill = section_fill
                row += 1

                hist_headers = ['Period', f'Avg Price ({unit})']
                for ci, h in enumerate(hist_headers, 1):
                    cell = ws.cell(row=row, column=ci, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                row += 1

                # Reverse so most recent is first
                for rec in reversed(historical):
                    ws.cell(row=row, column=1, value=rec['label']).border = thin_border
                    cell = ws.cell(row=row, column=2, value=rec['avg_price'])
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    row += 1

            # Auto-width columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'parramacro_commodity_forecasts_{today}.xlsx'
    )


# ── Power BI flat-table endpoints ─────────────────────────────────────────

def _build_powerbi_tables(data):
    """Flatten nested forecast data into Power BI star-schema tables."""
    groups = data.get('groups', {})
    time_ctx = data.get('time_context', {})
    meta = data.get('meta', {})
    labels = time_ctx.get('labels', [])
    label_types = time_ctx.get('label_types', [])
    year_end_labels = time_ctx.get('year_end_labels', [])
    fy_keys = ['FY'] + [f'FY{i+2}' for i in range(len(year_end_labels) - 1)]

    fact_prices = []
    dim_scenarios = []
    dim_commodities = []
    seen_scenarios = set()

    for group_name, group in groups.items():
        scenario_weights = group.get('scenario_weights', {})
        scenario_labels = group.get('scenario_labels', {})
        scenario_colors = group.get('scenario_colors', {})
        scenario_order = group.get('scenario_order', [])
        group_colors = group.get('colors', {})

        # Build dim_scenarios for this group
        for idx, sc in enumerate(scenario_order):
            key = (group_name, sc)
            if key not in seen_scenarios:
                seen_scenarios.add(key)
                dim_scenarios.append({
                    'Group': group_name,
                    'Scenario': sc,
                    'Weight': scenario_weights.get(sc),
                    'Description': scenario_labels.get(sc, ''),
                    'Color_Hex': scenario_colors.get(sc, ''),
                    'Sort_Order': idx,
                })

        for comm_name, info in group.get('commodities', {}).items():
            # dim_commodities
            dim_commodities.append({
                'Commodity': comm_name,
                'Group': group_name,
                'Unit': info.get('unit', ''),
                'Ticker': info.get('ticker', ''),
                'Latest_Close': info.get('latest_close'),
                'Color_Hex': group_colors.get(comm_name, ''),
            })

            # Historical rows → fact_prices (scenario = "Actual")
            for h in info.get('historical', []):
                year_val = h.get('year')
                quarter_val = h.get('quarter')
                fact_prices.append({
                    'Commodity': comm_name,
                    'Group': group_name,
                    'Scenario': 'Actual',
                    'Period': h.get('label', ''),
                    'Period_Type': 'historical',
                    'Year': year_val,
                    'Quarter': quarter_val,
                    'Price': h.get('avg_price'),
                    'Is_FY': False,
                })

            # Forecast rows → fact_prices (all scenarios)
            scenarios = info.get('scenarios', {})
            for sc_name, sc_data in scenarios.items():
                # Quarter-level rows
                for i, lbl in enumerate(labels):
                    price = sc_data.get(lbl)
                    if price is None:
                        continue
                    lt = label_types[i] if i < len(label_types) else 'forecast'
                    # Parse year/quarter from label
                    yr, qtr = _parse_period_label(lbl, time_ctx)
                    fact_prices.append({
                        'Commodity': comm_name,
                        'Group': group_name,
                        'Scenario': sc_name,
                        'Period': lbl,
                        'Period_Type': lt,
                        'Year': yr,
                        'Quarter': qtr,
                        'Price': price,
                        'Is_FY': False,
                    })

                # FY rows
                for fi, fy_key in enumerate(fy_keys):
                    fy_val = sc_data.get(fy_key)
                    if fy_val is None:
                        continue
                    fy_label = year_end_labels[fi] if fi < len(year_end_labels) else fy_key
                    # Extract year from "FY 2026"
                    fy_year = None
                    try:
                        fy_year = int(fy_label.split()[-1])
                    except (ValueError, IndexError):
                        pass
                    fact_prices.append({
                        'Commodity': comm_name,
                        'Group': group_name,
                        'Scenario': sc_name,
                        'Period': fy_label,
                        'Period_Type': 'fy',
                        'Year': fy_year,
                        'Quarter': None,
                        'Price': fy_val,
                        'Is_FY': True,
                    })

    return {
        'fact_prices': fact_prices,
        'dim_scenarios': dim_scenarios,
        'dim_commodities': dim_commodities,
        'meta': {
            'source': meta.get('source', 'ParraMacro'),
            'last_updated': meta.get('last_updated', ''),
            'method': meta.get('method', ''),
            'commodities_count': meta.get('commodities_count', 0),
            'fact_rows': len(fact_prices),
            'forecast_labels': labels,
            'year_end_labels': year_end_labels,
        },
    }


def _parse_period_label(lbl, time_ctx):
    """Extract (year, quarter) from labels like 'Q1*', 'Q2', \"Q1'27\", '2024 Q3'."""
    forecast_year = time_ctx.get('year', datetime.utcnow().year)
    # Historical: "2024 Q3"
    if ' Q' in lbl:
        parts = lbl.split(' Q')
        return int(parts[0]), int(parts[1])
    # Next-year shorthand: "Q1'27"
    if "'" in lbl:
        q_part = lbl.split("'")
        qtr = int(q_part[0].replace('Q', ''))
        yr = 2000 + int(q_part[1])
        return yr, qtr
    # Current year: "Q1*", "Q2", "Q3", "Q4"
    clean = lbl.replace('*', '').replace('Q', '')
    try:
        return forecast_year, int(clean)
    except ValueError:
        return None, None


@api_bp.route('/forecasts/powerbi')
def get_forecasts_powerbi():
    """Return flat Power BI-optimized tables (fact + dimensions) as JSON."""
    data = get_forecast_data()
    if not data:
        return jsonify({'error': 'Forecast data unavailable'}), 503
    return jsonify(_build_powerbi_tables(data))


@api_bp.route('/forecasts/powerbi/export')
def export_forecasts_powerbi_excel():
    """Generate flat Power BI-optimized Excel with Fact_Prices, Dim_Scenarios, Dim_Commodities sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_forecast_data()
    if not data:
        return jsonify({'error': 'Forecast data unavailable'}), 503

    tables = _build_powerbi_tables(data)
    wb = Workbook()

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    num_fmt = '#,##0.00'

    # ── Sheet 1: Fact_Prices ──
    ws = wb.active
    ws.title = 'Fact_Prices'
    fact_cols = ['Commodity', 'Group', 'Scenario', 'Period', 'Period_Type', 'Year', 'Quarter', 'Price', 'Is_FY']
    for ci, col_name in enumerate(fact_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for ri, row_data in enumerate(tables['fact_prices'], 2):
        for ci, col_name in enumerate(fact_cols, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if col_name == 'Price' and val is not None:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal='right')

    # ── Sheet 2: Dim_Scenarios ──
    ws2 = wb.create_sheet('Dim_Scenarios')
    scen_cols = ['Group', 'Scenario', 'Weight', 'Description', 'Color_Hex', 'Sort_Order']
    for ci, col_name in enumerate(scen_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for ri, row_data in enumerate(tables['dim_scenarios'], 2):
        for ci, col_name in enumerate(scen_cols, 1):
            val = row_data.get(col_name)
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = thin_border

    # ── Sheet 3: Dim_Commodities ──
    ws3 = wb.create_sheet('Dim_Commodities')
    comm_cols = ['Commodity', 'Group', 'Unit', 'Ticker', 'Latest_Close', 'Color_Hex']
    for ci, col_name in enumerate(comm_cols, 1):
        cell = ws3.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for ri, row_data in enumerate(tables['dim_commodities'], 2):
        for ci, col_name in enumerate(comm_cols, 1):
            val = row_data.get(col_name)
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if col_name == 'Latest_Close' and val is not None:
                cell.number_format = num_fmt

    # Auto-width all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 3, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'parramacro_powerbi_commodities_{today}.xlsx'
    )


@api_bp.route('/weo/<indicator>')
def get_weo(indicator):
    """Return IMF WEO data for the given indicator (cached 24 hours)."""
    data = get_weo_data(indicator)
    return jsonify(data)


@api_bp.route('/weo/<indicator>/export')
def export_weo_excel(indicator):
    """Generate Excel file with WEO indicator data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_weo_data(indicator)
    countries = data.get('countries', {})
    years = data.get('years', [])
    forecast_start = data.get('forecast_start_year')
    meta = data.get('meta', {})

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    forecast_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = meta.get('indicator_name', indicator)[:31]

    # Title
    ws.cell(row=1, column=1, value=meta.get('indicator_name', indicator))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=meta.get('source', 'IMF WEO'))
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='6B7280')

    # Headers
    headers = ['Country', 'ISO3'] + [str(y) for y in years]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        if forecast_start and col > 2:
            yr = years[col - 3]
            if yr >= forecast_start:
                cell.fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')

    # Data rows sorted by country name
    sorted_countries = sorted(countries.items(), key=lambda x: x[1].get('name', x[0]))
    for row_idx, (iso, cdata) in enumerate(sorted_countries, 5):
        ws.cell(row=row_idx, column=1, value=cdata.get('name', iso)).border = thin_border
        ws.cell(row=row_idx, column=2, value=iso).border = thin_border
        for i, yr in enumerate(years):
            val = cdata['values'].get(str(yr))
            cell = ws.cell(row=row_idx, column=3 + i, value=val)
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if forecast_start and yr >= forecast_start and val is not None:
                cell.fill = forecast_fill

    # Auto-width
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 8
    for i in range(len(years)):
        col_letter = ws.cell(row=4, column=3 + i).column_letter
        ws.column_dimensions[col_letter].width = 10
    ws.freeze_panes = 'C5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'weo_{indicator}_{today}.xlsx'
    )


@api_bp.route('/wb/<path:indicator>')
def get_world_bank(indicator):
    """Return World Bank data for the given indicator (cached 24 hours)."""
    data = get_wb_data(indicator)
    return jsonify(data)


@api_bp.route('/em-vulnerability')
def get_em_vulnerability():
    """Return EM external vulnerability bubble-chart dataset (cached 24 hours)."""
    data = get_em_vulnerability_data()
    return jsonify(data)


@api_bp.route('/em-vulnerability/missing')
def list_em_missing_st_debt():
    """Return which EMs are currently missing ST-debt data, grouped by
    whether they'd drop into the default Top 40 view. Drives the targeted
    data-source hunt for the holdouts.
    """
    data = get_em_vulnerability_data()
    countries = data.get('countries', {})
    missing_em = []
    missing_non_em = []
    for iso, r in countries.items():
        if r.get('reserves_to_st_debt_pct') is not None:
            continue
        row = {
            'iso3': iso,
            'name': r.get('name'),
            'gdp_usd': r.get('gdp_usd'),
            'em_rank': r.get('em_rank'),
            'gdp_rank': r.get('gdp_rank'),
            'basic_balance_pct_gdp': r.get('basic_balance_pct_gdp'),
            'reserves_usd': r.get('reserves_usd'),
            'reserves_source': r.get('reserves_source'),
            'st_debt_usd': r.get('st_debt_usd'),
            'st_debt_year': r.get('st_debt_year'),
            'st_debt_source': r.get('st_debt_source'),
        }
        if r.get('is_em'):
            missing_em.append(row)
        else:
            missing_non_em.append(row)
    missing_em.sort(key=lambda x: -(x.get('gdp_usd') or 0))
    missing_non_em.sort(key=lambda x: -(x.get('gdp_usd') or 0))
    return jsonify({
        'missing_em_count': len(missing_em),
        'missing_em': missing_em,
        'missing_non_em_count': len(missing_non_em),
        'missing_non_em_sample': missing_non_em[:20],
    })


@api_bp.route('/em-vulnerability/diagnose')
def diagnose_em_vulnerability():
    """Temporary diagnostic — probes WB API for the right QEDS source + indicator.

    Run this on prod (where api.worldbank.org is reachable) and paste the
    output; the dev sandbox is behind an allowlist that blocks WB, so this
    has to execute server-side. Remove once the right (source, indicator)
    pair is confirmed.
    """
    import requests

    result = {
        'step1_sources': None,
        'step2_indicators_with_short': {},
        'step3_chile_sample': {},
    }

    # Step 1: list all sources, look for anything with "Debt" or "QEDS".
    try:
        r = requests.get(
            'https://api.worldbank.org/v2/sources?format=json&per_page=100',
            timeout=30,
        )
        r.raise_for_status()
        doc = r.json()
        sources = doc[1] if isinstance(doc, list) and len(doc) > 1 else []
        result['step1_sources'] = [
            {'id': s.get('id'), 'name': s.get('name'), 'code': s.get('code')}
            for s in sources
            if 'debt' in (s.get('name', '') or '').lower()
            or 'qeds' in (s.get('name', '') or '').lower()
        ]
    except Exception as e:
        result['step1_sources'] = {'error': str(e)}

    # Step 2: for candidate source IDs, list short-term debt indicators.
    # Source 22 (QEDS SDDS) has many granular breakdowns; narrow to the ones
    # that look like the "all sectors, all instruments" aggregate.
    # Added source 54 = JEDH (Joint External Debt Hub) which combines
    # SDDS/GDDS/BIS/OECD and likely carries the cleanest aggregate.
    for source_id in ('6', '22', '23', '54', '81'):
        try:
            r = requests.get(
                f'https://api.worldbank.org/v2/sources/{source_id}/indicators'
                f'?format=json&per_page=5000',
                timeout=60,
            )
            r.raise_for_status()
            doc = r.json()
            indicators = doc[1] if isinstance(doc, list) and len(doc) > 1 else []
            # Only the aggregates: "All Sectors" + "Short-term" + (all
            # instruments OR no sub-qualifier). Exclude per-instrument
            # (Currency, Loans, Debt Securities, Trade) and per-sector
            # (Public, Private) breakdowns.
            aggregate_matches = []
            for i in indicators:
                if not isinstance(i, dict):
                    continue
                name = (i.get('name', '') or '').lower()
                if 'short-term' not in name:
                    continue
                if 'all sectors' not in name:
                    continue
                # Avoid partial-instrument or sector-specific ones
                skips = ['currency', 'loan', 'trade credit', 'debt securities',
                         'public sector', 'private sector', 'central bank',
                         'deposit-taking', 'beginning', 'market value',
                         'nominal value', 'diff.', 'other sector']
                if any(s in name for s in skips):
                    continue
                aggregate_matches.append({
                    'id': i.get('id'), 'name': i.get('name', '')[:120],
                })
            result['step2_indicators_with_short'][source_id] = aggregate_matches[:20]
        except Exception as e:
            result['step2_indicators_with_short'][source_id] = {'error': str(e)}

    # Step 3: probe the aggregate "all sectors, short-term, all instruments"
    # across QEDS SDDS, QEDS GDDS, and JEDH. For each, report how many
    # records have *non-null values* for each of the stranded EMs and the
    # most recent period+value found.
    result['step3_chile_sample'] = {}
    STRANDED_EMS = ['ARE', 'QAT', 'GTM', 'IRN', 'ISR', 'KWT', 'BHR', 'OMN',
                    'CHL', 'LBN', 'LBY']
    for source_id, ind in (
        # QEDS SDDS — correct aggregate
        ('22', 'DT.DOD.DSTC.CD.US'),
        # QEDS GDDS — correct aggregate
        ('23', 'DT.DOD.DECT.CD.ST.US'),
        # JEDH — try both known aggregates + WDI code
        ('54', 'DT.DOD.DSTC.CD'),
        ('54', 'DT.DOD.DSTC.CD.US'),
        ('54', 'DT.DOD.DECT.CD.ST.US'),
        # IDS DSSI — last resort for low-income / small-state EMs
        ('81', 'DT.DOD.DSTC.CD'),
        # QEDS SDDS granular (debt securities only) as partial fallback
        ('22', 'DT.DOD.DECT.CD.ST.TD.NV.US'),
    ):
        try:
            r = requests.get(
                f'https://api.worldbank.org/v2/country/all/indicator/{ind}'
                f'?format=json&mrv=20&source={source_id}&per_page=20000',
                timeout=60,
            )
            doc = r.json() if r.ok else None
            records = doc[1] if isinstance(doc, list) and len(doc) > 1 and doc[1] else []
            # Count records with a non-null value
            populated = [rec for rec in records
                         if isinstance(rec, dict) and rec.get('value') is not None]
            populated_isos = {
                (rec.get('country') or {}).get('id') for rec in populated
            }
            per_country = {}
            for iso in STRANDED_EMS:
                iso_records = [rec for rec in populated
                               if (rec.get('country') or {}).get('id') == iso]
                if iso_records:
                    # Most recent period+value
                    latest = sorted(iso_records,
                                    key=lambda r: r.get('date', ''),
                                    reverse=True)[0]
                    per_country[iso] = {
                        'period': latest.get('date'),
                        'value': latest.get('value'),
                    }
                else:
                    per_country[iso] = None
            result['step3_chile_sample'][f'{source_id}/{ind}'] = {
                'http': r.status_code,
                'total_records': len(records),
                'populated_records': len(populated),
                'unique_populated_isos': len(populated_isos),
                'populated_iso_sample': sorted(iso for iso in populated_isos if iso)[:20],
                'stranded_em_coverage': per_country,
            }
        except Exception as e:
            result['step3_chile_sample'][f'{source_id}/{ind}'] = {'error': str(e)}

    # Step 4: Enumerate IMF datasets on DBnomics for external debt.
    # Source 54 (JEDH) may or may not cover the stranded EMs — if not, a
    # direct IMF external debt dataset might.
    result['step4_dbnomics_imf'] = {}
    try:
        r = requests.get(
            'https://api.db.nomics.world/v22/providers/IMF',
            timeout=30,
        )
        if r.ok:
            doc = r.json()
            ds_list = (doc.get('category_tree') or [])
            flat = []
            def _walk(nodes):
                for n in nodes or []:
                    if n.get('code'):
                        flat.append((n.get('code'), n.get('name', '')))
                    _walk(n.get('children') or [])
            _walk(ds_list)
            debt_like = [(c, n) for c, n in flat
                         if 'debt' in (n or '').lower()
                         or 'sdds' in (n or '').lower()
                         or 'external' in (n or '').lower()]
            result['step4_dbnomics_imf']['imf_datasets_debt_related'] = debt_like[:30]
    except Exception as e:
        result['step4_dbnomics_imf']['enumerate_error'] = str(e)

    return jsonify(result)


@api_bp.route('/em-vulnerability/export')
def export_em_vulnerability_excel():
    """Generate Excel file with EM external vulnerability metrics.

    Two sheets:
      1. "Bubble Chart" — clean X / Y / Size / Label layout with an
         embedded Excel bubble chart already configured. Users can re-create
         or restyle the chart natively without touching column references.
      2. "Full Metrics" — all per-country fields including data-source
         provenance for the current account series.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BubbleChart, Reference, Series

    data = get_em_vulnerability_data()
    countries = data.get('countries', {})
    meta = data.get('meta', {})

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    em_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # Only countries that can be plotted (have basic balance). ST-debt-missing
    # ones still get a row but with blank X so Excel skips them — user can
    # tell from the blank cell why they're not on the chart.
    plottable = [
        r for r in countries.values()
        if r.get('basic_balance_pct_gdp') is not None
    ]
    em_rows = sorted(
        [r for r in plottable if r.get('is_em')],
        key=lambda r: -(r.get('gdp_usd') or 0),
    )
    dm_rows = sorted(
        [r for r in plottable if not r.get('is_em')],
        key=lambda r: -(r.get('gdp_usd') or 0),
    )
    chart_rows = em_rows + dm_rows  # EM first so they're a contiguous series

    wb = Workbook()

    # ── Sheet 1: Bubble Chart (chart-data layout + embedded chart) ──────────
    chart_ws = wb.active
    chart_ws.title = 'Bubble Chart'

    chart_ws.cell(row=1, column=1, value='EM External Vulnerability — Bubble Chart Data')
    chart_ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    chart_ws.cell(row=2, column=1, value=meta.get('source', 'World Bank'))
    chart_ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='6B7280')
    chart_ws.cell(
        row=3, column=1,
        value=(
            'X = Reserves / Short-Term External Debt (%)  ·  '
            'Y = Basic Balance (Current Account + Net FDI, % GDP)  ·  '
            'Bubble size = Nominal GDP ($B). '
            'EM rows (red shading) are listed first; the embedded chart '
            'splits them into two series for color coding.'
        ),
    )
    chart_ws.cell(row=3, column=1).font = Font(italic=True, size=9, color='6B7280')

    chart_headers = [
        'Country', 'ISO3', 'EM',
        'X · Reserves / ST Debt (%)',
        'Y · Basic Balance (% GDP)',
        'Size · GDP ($B)',
        'Year', 'CA Source',
    ]
    HEADER_ROW = 5
    for col, h in enumerate(chart_headers, 1):
        cell = chart_ws.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    DATA_START = HEADER_ROW + 1
    for i, r in enumerate(chart_rows):
        row = DATA_START + i
        values = [
            r.get('name', ''),
            r.get('iso3', ''),
            'Yes' if r.get('is_em') else 'No',
            r.get('reserves_to_st_debt_pct'),  # X — may be None
            r.get('basic_balance_pct_gdp'),     # Y
            (r.get('gdp_usd') or 0) / 1e9,      # Size in $B for legibility
            r.get('year', ''),
            r.get('ca_source', ''),
        ]
        for col, v in enumerate(values, 1):
            cell = chart_ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if r.get('is_em'):
                cell.fill = em_fill
            if col in (4, 5):
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 6:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col == 7:
                cell.alignment = Alignment(horizontal='center')

    chart_ws.column_dimensions['A'].width = 28
    chart_ws.column_dimensions['B'].width = 7
    chart_ws.column_dimensions['C'].width = 6
    chart_ws.column_dimensions['D'].width = 18
    chart_ws.column_dimensions['E'].width = 18
    chart_ws.column_dimensions['F'].width = 14
    chart_ws.column_dimensions['G'].width = 7
    chart_ws.column_dimensions['H'].width = 18
    chart_ws.row_dimensions[HEADER_ROW].height = 32
    chart_ws.freeze_panes = f'A{DATA_START}'

    # ── Embedded bubble chart ─────────────────────────────────────────────
    if chart_rows:
        em_count = len(em_rows)
        dm_count = len(dm_rows)
        em_end = DATA_START + em_count - 1
        dm_start = em_end + 1
        dm_end = dm_start + dm_count - 1

        bubble = BubbleChart()
        bubble.style = 18
        bubble.title = 'EM External Vulnerability'
        bubble.x_axis.title = 'Foreign Reserves / Short-Term External Debt (%)'
        bubble.y_axis.title = 'Basic Balance (Current Account + Net FDI, % GDP)'
        bubble.x_axis.scaling.min = 0
        bubble.height = 14
        bubble.width = 24
        bubble.legend.position = 'b'

        if em_count > 0:
            em_series = Series(
                values=Reference(chart_ws, min_col=5, min_row=DATA_START, max_row=em_end),
                xvalues=Reference(chart_ws, min_col=4, min_row=DATA_START, max_row=em_end),
                zvalues=Reference(chart_ws, min_col=6, min_row=DATA_START, max_row=em_end),
                title='Emerging Markets',
            )
            bubble.series.append(em_series)
        if dm_count > 0:
            dm_series = Series(
                values=Reference(chart_ws, min_col=5, min_row=dm_start, max_row=dm_end),
                xvalues=Reference(chart_ws, min_col=4, min_row=dm_start, max_row=dm_end),
                zvalues=Reference(chart_ws, min_col=6, min_row=dm_start, max_row=dm_end),
                title='Advanced / Other',
            )
            bubble.series.append(dm_series)

        # Anchor the chart to the right of the data, top-aligned with headers.
        chart_ws.add_chart(bubble, 'J5')

    # ── Sheet 2: Full Metrics (everything we have) ────────────────────────
    ws = wb.create_sheet(title='Full Metrics')

    ws.cell(row=1, column=1, value='EM External Vulnerability — Full Metrics')
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=meta.get('source', 'World Bank'))
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='6B7280')

    headers = [
        'Country', 'ISO3', 'EM', 'As-Of Year',
        'GDP (USD)', 'GDP Source', 'GDP Year', 'GDP Rank', 'EM Rank',
        'Current Account (% GDP)', 'CA Source', 'CA Year',
        'FDI Net (% GDP)', 'Basic Balance (% GDP)',
        'Reserves (USD)', 'Reserves Source', 'Reserves Period',
        'Short-Term External Debt (USD)', 'ST Debt Year', 'ST Debt Source',
        'Reserves / ST Debt (%)',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    sorted_rows = sorted(
        countries.values(),
        key=lambda r: (-(r.get('gdp_usd') or 0), r.get('name', '')),
    )
    for row_idx, r in enumerate(sorted_rows, 5):
        values = [
            r.get('name', ''),
            r.get('iso3', ''),
            'Yes' if r.get('is_em') else 'No',
            r.get('year', ''),
            r.get('gdp_usd'),
            r.get('gdp_source', ''),
            r.get('gdp_year', ''),
            r.get('gdp_rank'),
            r.get('em_rank'),
            r.get('ca_pct_gdp'),
            r.get('ca_source', ''),
            r.get('ca_year', ''),
            r.get('fdi_pct_gdp'),
            r.get('basic_balance_pct_gdp'),
            r.get('reserves_usd'),
            r.get('reserves_source', ''),
            r.get('reserves_period', ''),
            r.get('st_debt_usd'),
            r.get('st_debt_year', ''),
            r.get('st_debt_source', ''),
            r.get('reserves_to_st_debt_pct'),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            if col >= 4:
                cell.alignment = Alignment(horizontal='right')
                if col in (5, 15, 18):
                    cell.number_format = '#,##0'
                elif col in (10, 13, 14, 21):
                    cell.number_format = '0.00'

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['P'].width = 16
    ws.column_dimensions['Q'].width = 14
    ws.column_dimensions['S'].width = 10
    ws.column_dimensions['T'].width = 22
    for col_letter in ('E', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'R', 'U'):
        ws.column_dimensions[col_letter].width = 16
    ws.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'em_external_vulnerability_{today}.xlsx'
    )


@api_bp.route('/wb/<path:indicator>/export')
def export_wb_excel(indicator):
    """Generate Excel file with World Bank indicator data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_wb_data(indicator)
    countries = data.get('countries', {})
    years = data.get('years', [])
    meta = data.get('meta', {})

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = meta.get('indicator_name', indicator)[:31]

    # Title
    ws.cell(row=1, column=1, value=meta.get('indicator_name', indicator))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value='World Bank · World Development Indicators')
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='6B7280')

    # Headers
    headers = ['Country', 'ISO3'] + [str(y) for y in years]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows sorted by country name
    sorted_countries = sorted(countries.items(), key=lambda x: x[1].get('name', x[0]))
    for row_idx, (iso, cdata) in enumerate(sorted_countries, 5):
        ws.cell(row=row_idx, column=1, value=cdata.get('name', iso)).border = thin_border
        ws.cell(row=row_idx, column=2, value=iso).border = thin_border
        for i, yr in enumerate(years):
            val = cdata['values'].get(str(yr))
            cell = ws.cell(row=row_idx, column=3 + i, value=val)
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Auto-width
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 8
    for i in range(len(years)):
        col_letter = ws.cell(row=4, column=3 + i).column_letter
        ws.column_dimensions[col_letter].width = 10
    ws.freeze_panes = 'C5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = indicator.replace('.', '_').replace('/', '_')
    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'worldbank_{safe_name}_{today}.xlsx'
    )


@api_bp.route('/cofer/export')
def export_reserves_excel():
    """Generate Excel file with all reserves data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = get_cofer_data()
    years = data.get('years', [])
    countries = data.get('countries', [])

    wb = Workbook()

    # Sheet 1: Total Reserves
    ws1 = wb.active
    ws1.title = 'Total Reserves'
    _write_reserves_sheet(ws1, years, countries, 'total_reserves')

    # Sheet 2: FX Reserves
    ws2 = wb.create_sheet('FX Reserves')
    _write_reserves_sheet(ws2, years, countries, 'fx_reserves')

    # Sheet 3: Gold Reserves
    ws3 = wb.create_sheet('Gold Reserves')
    _write_reserves_sheet(ws3, years, countries, 'gold_reserves')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'central_bank_reserves_{today}.xlsx'
    )


def _write_reserves_sheet(ws, years, countries, field):
    """Write a reserves sheet with header + data rows."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')

    # Header
    headers = ['Country', 'ISO3'] + years
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, c in enumerate(countries, 2):
        ws.cell(row=row_idx, column=1, value=c['name'])
        ws.cell(row=row_idx, column=2, value=c['iso3'])
        values = c.get(field, [])
        for i, val in enumerate(values):
            if val is not None:
                ws.cell(row=row_idx, column=3 + i, value=val)

    # Auto-width first two columns
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 8
    for i in range(len(years)):
        col_letter = ws.cell(row=1, column=3 + i).column_letter
        ws.column_dimensions[col_letter].width = 12

    ws.freeze_panes = 'C2'


@api_bp.route('/substack')
def get_substack():
    """Return Substack posts from RSS feed (cached 1 hour)."""
    posts = get_substack_posts()
    return jsonify({'posts': posts})


@api_bp.route('/history')
def get_history():
    """Return daily historical snapshots of country scores (from SQLite).
    Optional query params: country=US, days=30
    """
    country_filter = request.args.get('country', '').upper()
    days_limit = int(request.args.get('days', 90))

    if country_filter:
        series = get_country_history(country_filter, days=days_limit)
        return jsonify({'country': country_filter, 'series': series})
    else:
        history = get_all_history(days=days_limit)
        sorted_dates = sorted(history.keys())
        return jsonify({'dates': sorted_dates, 'snapshots': history})


@api_bp.route('/history/<country_code>')
def get_country_history_endpoint(country_code):
    """Return full time series for a single country with all indicators."""
    code = country_code.upper()
    days = int(request.args.get('days', 90))
    series = get_country_history(code, days=days)
    if not series:
        return jsonify({'country': code, 'days': days, 'data_points': 0, 'series': []})
    return jsonify({
        'country': code,
        'days': days,
        'data_points': len(series),
        'series': series
    })


@api_bp.route('/anomalies')
def get_anomalies():
    """Return countries with significant score changes (>10 points)."""
    threshold = float(request.args.get('threshold', 10.0))
    anomalies = detect_anomalies(threshold_delta=threshold)
    return jsonify({'anomalies': anomalies})


@api_bp.route('/status')
def get_status():
    """Return system health info."""
    import os
    from backend.cache.database import get_history_dates
    last_refresh = store.get_last_refresh()
    scores_file_exists = os.path.exists(Config.SCORES_FILE)
    db_exists = os.path.exists(Config.DB_FILE)

    history_days = 0
    try:
        history_days = len(get_history_dates())
    except Exception:
        pass

    return jsonify({
        'status': 'ok',
        'last_refresh': last_refresh.isoformat() if last_refresh else None,
        'countries_tracked': store.country_count(),
        'hotspot_count': len(store.get_hotspots(Config.HOTSPOT_THRESHOLD)),
        'persistence': {
            'scores_file': scores_file_exists,
            'database': db_exists,
            'history_days': history_days,
            'total_score_rows': get_score_count(),
        }
    })


# ══════════════════════════════════════════════════════════════════════════════
# SOVEREIGN DEBT INDICATOR
# ══════════════════════════════════════════════════════════════════════════════

@api_bp.route('/sovereign-debt')
def get_sovereign_debt():
    """Return sovereign debt estimates for all countries."""
    data = get_sovereign_debt_data()
    return jsonify(data)


@api_bp.route('/sovereign-debt/export')
def export_sovereign_debt_excel():
    """Generate Excel file with sovereign debt indicator data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_sovereign_debt_data()
    countries = data.get('countries', {})

    if not countries:
        return jsonify({'error': 'No sovereign debt data available'}), 404

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    tier_fills = {
        'Critical': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
        'High': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'Elevated': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'Moderate': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'Low': PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sovereign Debt Indicator'

    # Title
    ws.cell(row=1, column=1, value='ParraMacro Sovereign Debt Indicator')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')
    ws.cell(row=2, column=1, value='Estimated Actual Debt Including Shadow/Hidden Components')
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color='6B7280')
    summary = data.get('summary', {})
    ws.cell(row=3, column=1,
            value=f'{summary.get("total_countries", 0)} countries  |  '
                  f'Avg official: {summary.get("avg_official", "N/A")}%  |  '
                  f'Avg estimated: {summary.get("avg_estimated", "N/A")}%  |  '
                  f'Avg gap: {summary.get("avg_gap", "N/A")}pp')
    ws.cell(row=3, column=1).font = Font(size=9, color='6B7280')

    # Headers
    headers = [
        'Country', 'ISO3', 'Region',
        'Official Debt (% GDP)', 'Est. Actual Debt (% GDP)', 'Debt Gap (pp)',
        'Floor (% GDP)', 'Ceiling (% GDP)',
        'GDP ($B)', 'External Debt ($B)', 'BIS Claims ($B)', 'Chinese Lending ($B)',
        'Governance Score', 'Risk Tier',
        'ST Debt ($B)', 'LT Debt ($B)', 'ST Share %',
        'Svc/Exports %', 'Int/Revenue %', 'Reserve Coverage %',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Data rows — sorted by estimated debt descending
    sorted_items = sorted(countries.items(),
                          key=lambda x: x[1].get('estimated_debt_gdp') or 0,
                          reverse=True)

    for row_idx, (iso3, c) in enumerate(sorted_items, 6):
        vals = [
            c.get('name', iso3),
            iso3,
            c.get('region', ''),
            c.get('official_debt_gdp'),
            c.get('estimated_debt_gdp'),
            c.get('debt_gap_pp'),
            c.get('confidence_floor_gdp'),
            c.get('confidence_ceiling_gdp'),
            c.get('gdp_usd_bn'),
            c.get('external_debt_usd_bn'),
            c.get('bis_claims_usd_bn'),
            c.get('chinese_lending_usd_bn'),
            c.get('wgi_avg'),
            c.get('risk_tier', ''),
            c.get('short_term_debt_usd_bn'),
            c.get('long_term_debt_usd_bn'),
            c.get('short_term_pct'),
            c.get('debt_service_pct_exports'),
            c.get('interest_pct_revenue'),
            c.get('reserve_coverage_pct'),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col >= 4 and col <= 13 and val is not None:
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal='center')

        # Color risk tier cell
        tier = c.get('risk_tier', '')
        tier_fill = tier_fills.get(tier)
        if tier_fill:
            ws.cell(row=row_idx, column=14).fill = tier_fill
            # Also highlight estimated debt column for Critical/High
            if tier in ('Critical', 'High'):
                ws.cell(row=row_idx, column=5).fill = tier_fill

    # Column widths
    widths = [22, 6, 20, 14, 16, 10, 10, 10, 10, 12, 12, 14, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = w

    ws.freeze_panes = 'A6'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'sovereign_debt_indicator_{today}.xlsx'
    )


# ══════════════════════════════════════════════════════════════════════════════
# FERTILIZER & EM INFLATION IMPACT
# ══════════════════════════════════════════════════════════════════════════════

@api_bp.route('/fertilizer-em-inflation')
def get_fertilizer_em_inflation():
    """Return fertilizer price forecasts and EM inflation impact estimates."""
    data = get_fertilizer_em_data()
    return jsonify(data)


# ══════════════════════════════════════════════════════════════════════════════
# YALE BUDGET LAB — AVERAGE EFFECTIVE TARIFF RATE
# ══════════════════════════════════════════════════════════════════════════════

@api_bp.route('/yale-tariff')
def get_yale_tariff():
    """Return Yale Budget Lab average effective US tariff rate series."""
    data = get_yale_tariff_data()
    return jsonify(data)


@api_bp.route('/yale-tariff/export')
def export_yale_tariff_excel():
    """Generate Excel file with the Yale Budget Lab tariff rate time series."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_yale_tariff_data()
    points = data.get('points', [])

    if not points:
        return jsonify({'error': 'No Yale tariff data available'}), 404

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    title_font = Font(bold=True, size=14, color='1E3A8A')
    label_font = Font(bold=True, size=11, color='1E3A8A')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    wb = Workbook()

    # ── Sheet 1: Time Series ────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Effective Tariff Rate'

    ws.cell(row=1, column=1, value='Yale Budget Lab — US Average Effective Tariff Rate')
    ws.cell(row=1, column=1).font = title_font
    ws.merge_cells('A1:D1')

    ws.cell(row=2, column=1, value=data.get('subtitle', ''))
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color='64748B')
    ws.merge_cells('A2:D2')

    headers = ['Date', 'Effective Tariff Rate (%)', 'Change vs. Prior (pp)', 'Policy Event']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    prior_value = None
    for i, p in enumerate(points):
        row = i + 5
        date = p.get('date', '')
        value = p.get('value')
        note = p.get('note', '')
        delta = (value - prior_value) if (value is not None and prior_value is not None) else None

        ws.cell(row=row, column=1, value=date).alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=row, column=2, value=value)
        c2.alignment = Alignment(horizontal='center')
        c2.number_format = '0.00'
        c3 = ws.cell(row=row, column=3, value=delta)
        c3.alignment = Alignment(horizontal='center')
        c3.number_format = '+0.00;-0.00;0.00'
        if delta is not None:
            if delta > 0:
                c3.font = Font(color='B91C1C')
            elif delta < 0:
                c3.font = Font(color='047857')
        c4 = ws.cell(row=row, column=4, value=note)
        c4.alignment = Alignment(horizontal='left', wrap_text=True)
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        prior_value = value

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 80

    # ── Sheet 2: Summary & Methodology ──────────────────────────────────
    ws2 = wb.create_sheet('Summary & Methodology')

    ws2.cell(row=1, column=1, value='Yale Budget Lab — Tariff Rate Tracker').font = title_font
    ws2.merge_cells('A1:B1')

    summary_rows = [
        ('Source', data.get('source', '')),
        ('Source URL', data.get('source_url', '')),
        ('Series', data.get('source_detail', '')),
        ('Unit', data.get('unit', '')),
        ('Frequency', data.get('frequency', '')),
        ('Last Updated', data.get('last_updated', '')),
        ('Latest Value', data.get('latest_value')),
        ('Latest Source', data.get('latest_source', '')),
    ]
    for i, (label, val) in enumerate(summary_rows):
        ws2.cell(row=3 + i, column=1, value=label).font = label_font
        ws2.cell(row=3 + i, column=2, value=val)

    # Peak / trough stats
    values = [p.get('value') for p in points if p.get('value') is not None]
    if values:
        peak = max(values)
        trough = min(values)
        peak_date = next(p.get('date') for p in points if p.get('value') == peak)
        trough_date = next(p.get('date') for p in points if p.get('value') == trough)
        post_baseline = points[1:] if len(points) > 1 else points
        pb_min = min(p.get('value') for p in post_baseline)
        pb_min_date = next(p.get('date') for p in post_baseline if p.get('value') == pb_min)

        stats_row = 3 + len(summary_rows) + 1
        ws2.cell(row=stats_row, column=1, value='Peak rate').font = label_font
        ws2.cell(row=stats_row, column=2, value=f'{peak:.2f}% on {peak_date}')
        ws2.cell(row=stats_row + 1, column=1, value='Trough (post-baseline)').font = label_font
        ws2.cell(row=stats_row + 1, column=2, value=f'{pb_min:.2f}% on {pb_min_date}')
        ws2.cell(row=stats_row + 2, column=1, value='Pre-trade-war baseline').font = label_font
        ws2.cell(row=stats_row + 2, column=2, value=f'{trough:.2f}% on {trough_date}')

    methodology = data.get('methodology', '')
    if methodology:
        meth_row = 3 + len(summary_rows) + 5
        ws2.cell(row=meth_row, column=1, value='Methodology').font = label_font
        ws2.cell(row=meth_row, column=2, value=methodology).alignment = Alignment(wrap_text=True, vertical='top')
        ws2.merge_cells(start_row=meth_row, start_column=2, end_row=meth_row + 6, end_column=4)
        ws2.row_dimensions[meth_row].height = 100

    notes = data.get('notes', [])
    if notes:
        notes_row = 3 + len(summary_rows) + 13
        ws2.cell(row=notes_row, column=1, value='Notes').font = label_font
        for i, note in enumerate(notes):
            ws2.cell(row=notes_row + i, column=2, value='• ' + note).alignment = Alignment(wrap_text=True, vertical='top')
            ws2.merge_cells(start_row=notes_row + i, start_column=2, end_row=notes_row + i, end_column=4)
            ws2.row_dimensions[notes_row + i].height = 45

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 60
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'yale_tariff_rates_{today}.xlsx'
    )


@api_bp.route('/fertilizer-em-inflation/export')
def export_fertilizer_em_inflation_excel():
    """Generate Excel file with fertilizer forecast and EM inflation impact data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_fertilizer_em_data()
    countries = data.get('countries', {})

    if not countries:
        return jsonify({'error': 'No fertilizer/EM inflation data available'}), 404

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    tier_fills = {
        'Tier 1': PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid'),
        'Tier 2': PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid'),
        'Tier 3': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'Tier 4': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
    }

    wb = Workbook()

    # ── Sheet 1: EM Inflation Impact ─────────────────────────────────────
    ws = wb.active
    ws.title = 'EM Inflation Impact'

    ws.cell(row=1, column=1, value='ParraMacro — Fertilizer & EM Inflation Impact')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')

    headers = ['Rank', 'Country', 'Region', 'Impact Tier',
               'Energy Impact (pp)', 'Food/Fert Impact (pp)', 'FX Multiplier',
               'Total Add\'l CPI (pp)', 'Current CPI (%)', 'New Est. CPI (%)',
               'Food CPI Weight', 'Fert Import Dep.']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sort by total impact descending
    sorted_countries = sorted(countries.items(),
                              key=lambda x: x[1].get('total_addl_cpi_pp', 0),
                              reverse=True)

    for rank, (name, c) in enumerate(sorted_countries, 1):
        row = rank + 3
        tier_key = c.get('impact_tier', '').split(' — ')[0]
        fill = tier_fills.get(tier_key)

        vals = [rank, name, c.get('region', ''), c.get('impact_tier', ''),
                c.get('energy_impact_pp'), c.get('food_fert_impact_pp'),
                c.get('fx_multiplier'), c.get('total_addl_cpi_pp'),
                c.get('current_cpi'), c.get('new_est_cpi'),
                c.get('food_cpi_wt'), c.get('fert_import_dep')]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if col >= 4 else 'left')

    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['D'].width = 26

    # ── Sheet 2: Fertilizer Forecasts ────────────────────────────────────
    ws2 = wb.create_sheet('Fertilizer Forecasts')
    ws2.cell(row=1, column=1, value='Fertilizer Price Forecasts ($/ton)')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')

    row_num = 3
    fert_data = data.get('fertilizer_forecasts', {})
    for fert_name in ['Urea', 'DAP', 'Potash']:
        fert = fert_data.get(fert_name, {})
        ws2.cell(row=row_num, column=1, value=fert_name)
        ws2.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1

        fert_headers = ['Scenario', 'Q1', 'Q2', 'Q3', 'Q4', 'FY 2026', 'YoY vs 2025']
        for c, h in enumerate(fert_headers, 1):
            cell = ws2.cell(row=row_num, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row_num += 1

        scenarios = fert.get('scenarios', {})
        for scenario_name in ['Base Case', 'Severe Case', 'Worst Case', 'Weighted Avg']:
            prices = scenarios.get(scenario_name, {})
            ws2.cell(row=row_num, column=1, value=scenario_name)
            ws2.cell(row=row_num, column=2, value=prices.get('Q1'))
            ws2.cell(row=row_num, column=3, value=prices.get('Q2'))
            ws2.cell(row=row_num, column=4, value=prices.get('Q3'))
            ws2.cell(row=row_num, column=5, value=prices.get('Q4'))
            ws2.cell(row=row_num, column=6, value=prices.get('FY 2026'))
            if scenario_name == 'Weighted Avg':
                yoy = fert.get('yoy_pct')
                if yoy:
                    ws2.cell(row=row_num, column=7, value=f'+{yoy}%')
            row_num += 1
        row_num += 1

    for col in range(1, 8):
        ws2.column_dimensions[chr(64 + col)].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'fertilizer_em_inflation_{today}.xlsx'
    )


# ══════════════════════════════════════════════════════════════════════════════
# INSURANCE / REINSURANCE INFLATION (VERIFIED @AIG.COM OR ADMIN GRANT)
# ══════════════════════════════════════════════════════════════════════════════

@api_bp.route('/insurance-inflation')
@insurance_access_required
def get_insurance_inflation():
    """Return insurance/reinsurance inflation data (verified @aig.com or admin grant)."""
    data = get_insurance_inflation_data()
    return jsonify(data)


@api_bp.route('/insurance-inflation/export')
@insurance_access_required
def export_insurance_inflation_excel():
    """Generate Excel file with insurance inflation data (verified @aig.com or admin grant)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    from backend.data_sources.insurance_inflation import compute_qoq, aggregate_monthly_to_quarterly

    freq = request.args.get('freq', 'quarterly')       # 'monthly' | 'quarterly'
    comparison = request.args.get('comparison', 'yoy')  # 'yoy' | 'qoq'
    region = request.args.get('region', 'all')           # 'all' | 'us' | 'uk' | 'eu'

    data = get_insurance_inflation_data()
    series = data.get('series', {})
    series_raw = data.get('series_raw', {})
    categories = data.get('categories', {})
    meta = data.get('series_meta', {})

    if not series:
        return jsonify({'error': 'No insurance inflation data available'}), 404

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    comp_label = 'QoQ % Change' if comparison == 'qoq' else 'YoY % Change'
    freq_label = 'Quarterly' if freq == 'quarterly' else 'Monthly'

    wb = Workbook()
    first = True

    for cat_key, cat_info in categories.items():
        cat_series = cat_info.get('series', [])
        if not cat_series:
            continue

        # Filter to series that actually have data + region filter
        active_series = [(k, meta.get(k, {})) for k in cat_series if series.get(k)]
        if region != 'all':
            prefix_map = {'us': ('us_',), 'uk': ('uk_',), 'eu': ('eu_', 'nl_', 'it_')}
            prefixes = prefix_map.get(region, ())
            active_series = [(k, m) for k, m in active_series if any(k.startswith(p) for p in prefixes)]
        if not active_series:
            continue

        if first:
            ws = wb.active
            ws.title = cat_info['label']
            first = False
        else:
            ws = wb.create_sheet(cat_info['label'])

        # Title
        ws.cell(row=1, column=1, value=f'{cat_info["label"]} — Insurance Inflation Indicators')
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')
        ws.cell(row=2, column=1, value=f'{comp_label}, {freq_label}. Source: ONS, Eurostat. Auto-refreshes every 24h.')
        ws.cell(row=2, column=1).font = Font(italic=True, size=10, color='6B7280')

        # Transform data based on freq/comparison params
        transformed = {}
        for s_key, s_meta in active_series:
            is_quarterly = s_meta.get('freq') == 'Q'

            if comparison == 'qoq':
                raw = series_raw.get(s_key, [])
                if raw:
                    transformed[s_key] = compute_qoq(raw, is_quarterly)
                else:
                    transformed[s_key] = []
            else:
                pts = series.get(s_key, [])
                if freq == 'quarterly' and not is_quarterly:
                    transformed[s_key] = aggregate_monthly_to_quarterly(pts)
                else:
                    transformed[s_key] = pts

        # Collect all unique dates
        all_dates = set()
        for s_key, _ in active_series:
            for pt in transformed.get(s_key, []):
                all_dates.add(pt['date'])
        sorted_dates = sorted(all_dates)

        # Build lookup: {series_key: {date: value}}
        lookups = {}
        for s_key, _ in active_series:
            lookups[s_key] = {pt['date']: pt['value'] for pt in transformed.get(s_key, [])}

        # Header row: Date | Series1 | Series2 | ...
        row_num = 4
        ws.cell(row=row_num, column=1, value='Date').font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        for col, (s_key, s_meta) in enumerate(active_series, 2):
            approx = ' *' if s_meta.get('approximate') else ''
            cell = ws.cell(row=row_num, column=col, value=f'{s_meta.get("label", s_key)}{approx}')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Source row
        row_num += 1
        ws.cell(row=row_num, column=1, value='Source').font = Font(italic=True, size=9, color='6B7280')
        for col, (s_key, s_meta) in enumerate(active_series, 2):
            display_freq = 'Q' if freq == 'quarterly' or comparison == 'qoq' else s_meta.get('freq', 'M')
            ws.cell(row=row_num, column=col, value=f'{s_meta.get("source", "")} ({display_freq})').font = Font(italic=True, size=9, color='6B7280')

        # Data rows: one row per date, all series as columns
        row_num += 1
        for i, date_str in enumerate(sorted_dates):
            ws.cell(row=row_num, column=1, value=date_str)
            fill = alt_fill if i % 2 == 0 else None
            if fill:
                ws.cell(row=row_num, column=1).fill = fill
            for col, (s_key, _) in enumerate(active_series, 2):
                val = lookups[s_key].get(date_str)
                cell = ws.cell(row=row_num, column=col)
                if val is not None:
                    cell.value = val
                    cell.number_format = '0.00'
                if fill:
                    cell.fill = fill
            row_num += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        from openpyxl.utils import get_column_letter
        for col in range(2, len(active_series) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 22

        # Footnote for approximate series
        has_approx = any(m.get('approximate') for _, m in active_series)
        if has_approx:
            row_num += 1
            ws.cell(row=row_num, column=1, value='* Approximate proxy (nearest available COICOP code)').font = Font(italic=True, size=9, color='6B7280')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.utcnow().strftime('%Y-%m-%d')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'insurance_inflation_{region}_{freq}_{comparison}_{today}.xlsx'
    )
