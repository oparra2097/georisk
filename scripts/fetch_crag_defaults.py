"""
Fetch the Bank of Canada Credit Rating Assessment Group (CRAG) database
of sovereign default events and write it into
``data/sovereign_defaults.csv`` in the schema this project expects.

The CRAG file is the authoritative public panel of sovereign default
events: ~1,300 country-year-instrument observations 1960–present,
covering bond defaults, bank-loan restructurings, Paris Club and London
Club agreements, and arrears. Updated annually (typically June). Free.

Usage::

    python scripts/fetch_crag_defaults.py
    # By default writes to data/sovereign_defaults.csv (overwrites).
    # --dry-run to inspect counts without writing.

If your network can't reach bankofcanada.ca, download the spreadsheet
manually and pass ``--input /path/to/file.xlsx``.
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import sys
import urllib.request
from pathlib import Path
from typing import List, Optional

# Allow running as `python scripts/fetch_crag_defaults.py` from repo root.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# Direct download link for the BoC-BoE Sovereign Default Database. The
# Bank of Canada re-hosts it under a year-stamped path on each annual
# refresh; bump this to the latest edition when it 404s.
CRAG_URL = (
    'https://www.bankofcanada.ca/wp-content/uploads/'
    '2025/10/BoC-BoE-Database-2025.xlsx'
)

# 2024+ workbooks ditched the wide instrument-string matrix and replaced
# it with a "by Issuer" block layout: each country gets one header row
# (integer index in col B, name in col C, data-score in col D) followed
# by ~7 creditor-type sub-rows whose year cells hold US$ amounts in
# default. This map buckets each creditor into the (event_type,
# instrument) tuple our schema expects. Multilateral arrears (IMF/IBRD/
# IDA/IADB) are intentionally skipped — they sit outside the credit-
# event taxonomy in defaults.py.
CREDITOR_MAP = {
    'Paris Club':              ('paris_club',    'paris_club'),
    'China':                   ('paris_club',    'paris_club'),
    'Other official creditors':('paris_club',    'paris_club'),
    'FC bank loans':           ('restructuring', 'bank_loan'),
    'FC bonds':                ('default',       'external_bond'),
    'Other private creditors': ('restructuring', 'bank_loan'),
    'LC debt':                 ('default',       'domestic'),
    'Domestic arrears':        ('arrears',       'domestic'),
}

OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent / 'data' / 'sovereign_defaults.csv'
)


def download_crag(url: str = CRAG_URL) -> bytes:
    """Download the CRAG xlsx; raise if the network call fails."""
    req = urllib.request.Request(
        url, headers={'User-Agent': 'parra-macro/1.0 (+https://parramacro.com)'}
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_crag_xlsx(blob: bytes) -> List[dict]:
    """Parse the BoC-BoE Sovereign Default Database xlsx into events.

    The 2024+ workbooks use a "by Issuer" block layout. Each country has
    one header row (int index in col B, name in col C, data-score in
    col D) followed by creditor sub-rows whose cells along the year
    columns hold US$ amounts in default for that creditor. We walk each
    creditor sub-row, group contiguous in-default years into a run, and
    emit one event per (iso3, creditor, run) using ``CREDITOR_MAP`` to
    bucket the creditor into our (event_type, instrument) schema.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError('openpyxl required: pip install openpyxl')

    try:
        from backend.data_sources.country_codes import get_iso3_for_name
    except ImportError:
        # No name->ISO3 helper exists in the project today — fall back to
        # the curated map below. Kept as a soft import so a future helper
        # can be picked up automatically.
        get_iso3_for_name = None

    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)

    # Pick the first sheet whose row 0 contains year integers — the
    # primary panel sheet is named "<year> in Progress" and shifts each
    # release.
    ws = None
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        first_row = next(candidate.iter_rows(values_only=True), None)
        if first_row and any(
            isinstance(v, int) and 1900 < v < 2100 for v in first_row
        ):
            ws = candidate
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError('CRAG sheet appears empty')

    year_cols: List[tuple] = []  # (col_idx, year)
    for j, v in enumerate(rows[0]):
        if isinstance(v, int) and 1900 < v < 2100:
            year_cols.append((j, v))
    if not year_cols:
        raise RuntimeError('Could not identify year columns in CRAG sheet')
    last_year = year_cols[-1][1]

    out: List[dict] = []
    current_iso3: Optional[str] = None
    current_country: Optional[str] = None

    for r in rows:
        if not r:
            continue
        b = r[1] if len(r) > 1 else None
        c = r[2] if len(r) > 2 else None
        d = r[3] if len(r) > 3 else None

        # Country header row.
        if (
            isinstance(b, int) and 0 < b < 1000
            and isinstance(c, str) and c.strip()
            and not c.lstrip().startswith('-')
            and isinstance(d, (int, float))
        ):
            current_country = c.strip()
            current_iso3 = _name_to_iso3(current_country, get_iso3_for_name)
            continue

        # Creditor sub-row.
        if not (current_iso3 and isinstance(c, str) and c.lstrip().startswith('-')):
            continue
        creditor = c.strip().lstrip('-').strip()
        bucket = CREDITOR_MAP.get(creditor)
        if bucket is None:
            continue
        event_type, instrument = bucket

        run_start: Optional[int] = None
        run_end: Optional[int] = None
        for j, year in year_cols:
            val = r[j] if j < len(r) else None
            if _is_in_default(val):
                if run_start is None:
                    run_start = year
                run_end = year
            elif run_start is not None:
                out.append(_event(
                    current_iso3, run_start, run_end,
                    event_type, instrument, creditor,
                ))
                run_start = None
                run_end = None
        if run_start is not None:
            # Trailing run: if it reaches the last covered year, treat as
            # ongoing (end_year=None) so the model flags it as live.
            end_out = None if run_end == last_year else run_end
            out.append(_event(
                current_iso3, run_start, end_out,
                event_type, instrument, creditor,
            ))

    return out


def _is_in_default(val) -> bool:
    """A BoC cell counts as 'in default' if it's a positive amount or
    the '****' sentinel (default present, amount unavailable)."""
    if val is None:
        return False
    if isinstance(val, str):
        s = val.strip()
        return bool(s) and s != '0'
    try:
        return float(val) > 0
    except (TypeError, ValueError):
        return False


def _event(iso3: str, start: int, end: Optional[int],
           event_type: str, instrument: str, creditor: str) -> dict:
    return {
        'iso3': iso3, 'start_year': start, 'end_year': end,
        'event_type': event_type, 'instrument': instrument,
        'source': 'CRAG', 'notes': f'BoC creditor: {creditor}',
    }


_FALLBACK_NAME_MAP = {
    # Spellings used by the BoC-BoE workbook (covers all 165 issuers in
    # the 2025 edition) plus a few legacy aliases.
    'Afghanistan': 'AFG', 'Albania': 'ALB', 'Algeria': 'DZA',
    'Angola': 'AGO', 'Anguila': 'AIA', 'Anguilla': 'AIA',
    'Antigua and Barbuda': 'ATG', 'Argentina': 'ARG', 'Armenia': 'ARM',
    'Aruba': 'ABW', 'Azerbaijan': 'AZE',
    'Bahamas': 'BHS', 'Bahamas, The': 'BHS',
    'Bangladesh': 'BGD', 'Barbados': 'BRB', 'Belarus': 'BLR',
    'Belize': 'BLZ', 'Benin': 'BEN', 'Bhutan': 'BTN', 'Bolivia': 'BOL',
    'Bosnia & Herzegovina': 'BIH', 'Bosnia and Herzegovina': 'BIH',
    'Botswana': 'BWA', 'Brazil': 'BRA', 'Bulgaria': 'BGR',
    'Burkina Faso': 'BFA', 'Burundi': 'BDI',
    'Cabo Verde': 'CPV', 'Cape Verde': 'CPV',
    'Cambodia': 'KHM', 'Cameroon': 'CMR',
    'Central African Republic': 'CAF', 'Chad': 'TCD', 'Chile': 'CHL',
    'China': 'CHN', 'Colombia': 'COL', 'Comoros': 'COM',
    'Rep. Of Congo (Brazzaville)': 'COG', 'Congo, Rep.': 'COG',
    'Dem. Rep. of Congo (Kinshasa)': 'COD', 'Congo, Dem. Rep.': 'COD',
    'Cook Islands': 'COK', 'Costa Rica': 'CRI',
    "Côte d’Ivoire": 'CIV', "Cote d'Ivoire": 'CIV', 'Cote d Ivoire': 'CIV',
    'Ivory Coast': 'CIV',
    'Croatia': 'HRV', 'Cuba': 'CUB',
    'Curaçao': 'CUW', 'Curacao': 'CUW',
    'Cyprus': 'CYP', 'Czechoslovakia': 'CZE', 'Czech Republic': 'CZE',
    'Djibouti': 'DJI', 'Dominica': 'DMA', 'Dominican Republic': 'DOM',
    'Ecuador': 'ECU', 'Egypt': 'EGY', 'Egypt, Arab Rep.': 'EGY',
    'El Salvador': 'SLV', 'Equatorial Guinea': 'GNQ', 'Eritrea': 'ERI',
    'eSwatini (Swaziland)': 'SWZ', 'Eswatini': 'SWZ', 'Swaziland': 'SWZ',
    'Ethiopia': 'ETH', 'Fiji': 'FJI', 'Gabon': 'GAB',
    'The Gambia': 'GMB', 'Gambia, The': 'GMB',
    'Georgia': 'GEO', 'Ghana': 'GHA', 'Greece': 'GRC', 'Grenada': 'GRD',
    'Guatemala': 'GTM', 'Guinea': 'GIN', 'Guinea-Bissau': 'GNB',
    'Guyana': 'GUY', 'Haiti': 'HTI', 'Honduras': 'HND', 'Hungary': 'HUN',
    'India': 'IND', 'Indonesia': 'IDN', 'Iran': 'IRN',
    'Iran, Islamic Rep.': 'IRN', 'Iraq': 'IRQ', 'Ireland': 'IRL',
    'Jamaica': 'JAM', 'Jordan': 'JOR', 'Kazakhstan': 'KAZ', 'Kenya': 'KEN',
    "Korea, Democratic People's Republic of (North)": 'PRK',
    'Korea, Dem. Peoples Rep.': 'PRK', 'North Korea': 'PRK',
    'Korea, Rep.': 'KOR', 'South Korea': 'KOR',
    'Kosovo': 'XKX',
    'Kyrgyz Republic': 'KGZ', 'Kyrgyzstan': 'KGZ',
    'Laos': 'LAO', 'Lao PDR': 'LAO',
    'Latvia': 'LVA', 'Lebanon': 'LBN', 'Lesotho': 'LSO', 'Liberia': 'LBR',
    'Libya': 'LBY', 'Lithuania': 'LTU', 'North Macedonia': 'MKD',
    'Macedonia, FYR': 'MKD',
    'Madagascar': 'MDG', 'Malawi': 'MWI', 'Maldives': 'MDV', 'Mali': 'MLI',
    'Marshall Islands': 'MHL', 'Mauritania': 'MRT', 'Mauritius': 'MUS',
    'Mexico': 'MEX', 'Micronesia': 'FSM', 'Moldova': 'MDA',
    'Mongolia': 'MNG', 'Montenegro': 'MNE', 'Morocco': 'MAR',
    'Mozambique': 'MOZ', 'Myanmar': 'MMR', 'Namibia': 'NAM',
    'Nauru': 'NRU', 'Nepal': 'NPL',
    'Netherlands Antilles': 'ANT',
    'Nicaragua': 'NIC', 'Niger': 'NER', 'Nigeria': 'NGA',
    'Pakistan': 'PAK', 'Palau': 'PLW', 'Panama': 'PAN',
    'Papua New Guinea': 'PNG', 'Paraguay': 'PRY', 'Peru': 'PER',
    'Philippines': 'PHL', 'Poland': 'POL', 'Portugal': 'PRT',
    'Puerto Rico': 'PRI', 'Romania': 'ROU', 'Rwanda': 'RWA',
    'Russia': 'RUS', 'Russian Federation': 'RUS',
    'USSR/Russian Federation': 'RUS', 'USSR': 'RUS',
    'St. Kitts & Nevis': 'KNA', 'Saint Kitts and Nevis': 'KNA',
    'St. Lucia': 'LCA', 'Saint Lucia': 'LCA',
    'St. Vincent and the Grenadines': 'VCT',
    'Saint Vincent and the Grenadines': 'VCT',
    'Samoa': 'WSM',
    'São Tomé and Príncipe': 'STP', 'Sao Tome and Principe': 'STP',
    'Senegal': 'SEN', 'Serbia': 'SRB',
    'Serbia and Montenegro': 'SRB', 'Yugoslavia': 'SRB',
    'Seychelles': 'SYC', 'Sierra Leone': 'SLE',
    'Sint Maarten': 'SXM',
    'Slovak Republic': 'SVK', 'Slovakia': 'SVK',
    'Slovenia': 'SVN', 'Solomon Islands': 'SLB', 'Somalia': 'SOM',
    'South Africa': 'ZAF', 'South Sudan': 'SSD',
    'Sri Lanka': 'LKA', 'Sudan': 'SDN', 'Suriname': 'SUR',
    'Syria': 'SYR', 'Syrian Arab Republic': 'SYR',
    'Tajikistan': 'TJK', 'Tanzania': 'TZA', 'Thailand': 'THA',
    'Togo': 'TGO', 'Tonga': 'TON',
    'Trinidad & Tobago': 'TTO', 'Trinidad and Tobago': 'TTO',
    'Tunisia': 'TUN', 'Turkey': 'TUR', 'Turkmenistan': 'TKM',
    'Tuvalu': 'TUV', 'Uganda': 'UGA', 'Ukraine': 'UKR',
    'United Kingdom': 'GBR',
    'United States': 'USA', 'United States of America': 'USA',
    'Uruguay': 'URY', 'Uzbekistan': 'UZB', 'Vanuatu': 'VUT',
    'Venezuela': 'VEN', 'Venezuela, RB': 'VEN',
    'Vietnam': 'VNM', 'Viet Nam': 'VNM',
    'West Bank & Gaza': 'PSE', 'West Bank and Gaza': 'PSE',
    'Yemen': 'YEM', 'Yemen, Rep.': 'YEM',
    'Zambia': 'ZMB', 'Zimbabwe': 'ZWE',
}


def _name_to_iso3(name: str, helper) -> Optional[str]:
    if helper is not None:
        try:
            iso3 = helper(name)
            if iso3:
                return iso3
        except Exception:
            pass
    return _FALLBACK_NAME_MAP.get(name.strip())


def write_csv(events: List[dict], path: Path) -> None:
    fieldnames = ['iso3', 'start_year', 'end_year', 'event_type',
                  'instrument', 'source', 'notes']
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        f.write(
            '# Sovereign default / restructuring events from the Bank of Canada CRAG\n'
            '# database. Auto-generated by scripts/fetch_crag_defaults.py.\n'
            '# Schema matches what backend/credit_default/defaults.py expects.\n'
        )
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for ev in events:
            row = {k: (ev.get(k) if ev.get(k) is not None else '') for k in fieldnames}
            writer.writerow(row)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', help='Local CRAG xlsx path (skip network).')
    parser.add_argument('--output', default=str(OUTPUT_PATH),
                        help='Where to write the harmonized CSV.')
    parser.add_argument('--dry-run', action='store_true',
                        help='Parse + summarize without writing.')
    args = parser.parse_args()

    if args.input:
        with open(args.input, 'rb') as f:
            blob = f.read()
        print(f'[crag] using local file: {args.input}')
    else:
        try:
            print(f'[crag] downloading {CRAG_URL} …')
            blob = download_crag()
        except Exception as e:
            print(f'[crag] download failed: {e}')
            print('[crag] re-run with --input <path> after manual download.')
            return 2

    events = parse_crag_xlsx(blob)
    print(f'[crag] parsed {len(events)} events across {len(set(e["iso3"] for e in events))} countries')

    type_counts = {}
    for e in events:
        type_counts[e['event_type']] = type_counts.get(e['event_type'], 0) + 1
    print(f'[crag] type breakdown: {type_counts}')

    if args.dry_run:
        print('[crag] dry-run — not writing.')
        return 0

    out = Path(args.output)
    write_csv(events, out)
    print(f'[crag] wrote {out}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
